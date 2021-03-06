from __future__ import absolute_import
from __future__ import print_function
from __future__ import unicode_literals
from   rgclass import test_house_devices_static_info
# import itertools
import pprint
import subprocess

import wget
from datetime import datetime
# from selenium.webdriver.support.ui import WebDriverWait
import traceback
import pickle
# import global.py
# import time
# from  pprint import pprint
from openpyxl import Workbook
# import requests
# from bs4 import BeautifulSoup
# import requests
# from selenium import webdriver
# import urllib3
import re
import smtplib
import sys
from time import sleep
# from datetime import datetime
import datetime
import pexpect
from rgclass import Nvg599Class
# from rgclass import NON_DFS_CHANNELS
# from rgclass import DFS_CHANNELS
# from rgclass import  test_house_devices_static_info

NON_DFS_CHANNELS = {"36", "40", "44", "48", "149", "153", "157", "161", "165"}
DFS_CHANNELS = {"52", "56", "60", "64", "100", "104", "108", "112", "116", "132", "136", "140", "144"}
ALL_BAND5_CHANNELS = {36, 40, 44, 48, 52, 56, 60, 64, 100, 104, 108, 112,
                      116, 132, 136, 140, 144, 149, 153, 157, 161, 165}

# from rgclass import  NON_DFS_CHANNELS
# from rgclass import  DFS_CHANNELS
# Find the current channel used for 5G
# Check the 5G channel used. If none DFS , set to DFS and note the setting
# enter the command to simulate radar detection
# verify that the channel changes to a non DFS channel

# def execute_factory_reset(nvg_599_dut,rf,rfa, test_name):
#     test_status = "Pass"
#     rf.write('Test ' + test_name + '\n')
#     print('Test:' + test_name + '\n')

def dfs_channel_change(nvg_599_dut,rf, rfa, test_name, start_5g_channel, end_5g_channel, airties_name = "None" ):
    global NON_DFS_CHANNELS
    global DFS_CHANNELS
    test_status = "Pass"
    rf.write('Test ' + test_name + ' Airties device:' + airties_name + '\n')
    print('Test:' + test_name + '\n')
    session = nvg_599_dut.session
    home_link = session.find_element_by_link_text("Device")
    home_link.click()
    current_5g_channel = nvg_599_dut.get_ui_home_network_status_value("ui_channel_5g")
    # var_type = type(current_5g_channel)
    # print('type is:' + str(var_type))
    if current_5g_channel != start_5g_channel:
        nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, start_5g_channel)
        print('setting channel to :' + start_5g_channel)
    else:
        print('current channel:' + current_5g_channel + ' equals start_channel: ' + start_5g_channel)

    print('sleeping five minute to establish initial conditions')
    sleep(300)
    # we want the IP of the Airties we are going to ping
    print('sleeping five minute to establish initial conditions')
    print(' Airties: ' + airties_name + '\n')
    print(' Airties:-------------------------------------------------------\n')

    # nvg_599_dut.login_nvg_599_cli()
    ip_lan_info_dict = nvg_599_dut.get_rg_ip_lan_info_dict()
    # first translate the device name to a mac using the
    for x, y in ip_lan_info_dict.items():
        print(x, y)
    airties_ip = "0.0.0.0"

    for device_mac in ip_lan_info_dict:
        if airties_name == ip_lan_info_dict[device_mac]['Name']:
            airties_ip = ip_lan_info_dict[device_mac]['IP']
            print(' Airties ip:  ' + airties_ip + '\n')

    if airties_ip == "0.0.0.0":
        rf.write('    ' + 'Airties device not found in sh IP lan: Aborting \n')
        print('   Airties device not found in sh IP lan: Aborting \n')
        return "Fail"

    print(' Airties ip: ' + airties_ip + '\n')
    print(' This is before (1) \n')


def trigger_dfs_channel_change(nvg_599_dut,rf, rfa, test_name, airties_name = "None" ):
    global NON_DFS_CHANNELS
    global DFS_CHANNELS
    test_status = "Pass"
    rf.write('Test ' + test_name + ' Airties device:' + airties_name + '\n')
    print('Test:' + test_name + '\n')
    session = nvg_599_dut.session
    home_link = session.find_element_by_link_text("Device")
    home_link.click()

    current_5g_channel = nvg_599_dut.get_ui_home_network_status_value("ui_channel_5g")
    var_type = type(current_5g_channel)
    print('type is:' + str(var_type))

    # airties_cli_session = nvg_599_dut.login_4920("192.168.1.72")
    # exit()
    # nvg_599_dut.login_nvg_599_cli()

    print('current_5g_channel' + current_5g_channel + '\n')
    if current_5g_channel in DFS_CHANNELS:
        result = "Current 5G:" + current_5g_channel + " is a DFS channel"
        result_str = str(result)
        rf.write('    ' + result_str + '\n')
        print('this is a DFS channel:' + str(current_5g_channel))
    else:
        print('this is a non DFS channel,  Changing to DFS channel 100')
        # def ui_set_bw_channel(self, band, bandwidth, channel):
        # dfs_results_file.write("Changing to DFS channel 100, bandwidth 80\n")
        rf.write('    ' + 'Changing to DFS channel 100, bandwidth 80\n')
        print(' Airties before 5 \n')
        # new_session = nvg_599_dut.login_4920("192.168.1.67")
        # new_session.sendline('\n')
        # new_session.expect('#')
        # new_session.close()
        nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, 100)
        print('setting channel to DFS channel 100')
    sleep(220)
    # nvg_599_dut.login_4920("192.168.1.72")
    # this is the question, are we trggering dfs from the airties or from the RG
    # this is the default, we are triggering dfs from the RG 5 g side

    if airties_name == "None":
        print(' Airties:' + airties_name + '\n')
        print('triggering DFS from the RG \n')
        nvg_599_dut.login_nvg_599_cli()
        nvg_599_dut.telnet_cli_session.sendline()
        nvg_599_dut.telnet_cli_session.expect(">")
        # this is the IP used for build prior to corvus3 d13/d11
        # nvg_599_dut.telnet_cli_session.sendline("telnet 192.168.1.1")
        nvg_599_dut.telnet_cli_session.sendline("telnet 203.0.113.2")
        nvg_599_dut.telnet_cli_session.expect("#")
        # I think the 2 is to trgger radar on the 4920
        # nvg_599_dut.telnet_cli_session.sendline("wl -i eth1 radar 2")
        nvg_599_dut.telnet_cli_session.sendline("wl -i eth1 radar 2")
        sleep(10)
        nvg_599_dut.telnet_cli_session.close()

    else:
        print(' Airties: ' + airties_name + '\n')
        print(' Airties:-------------------------------------------------------\n')

        # nvg_599_dut.login_nvg_599_cli()
        ip_lan_info_dict = nvg_599_dut.get_rg_ip_lan_info_dict()
        # first translate the device name to a mac using the
        for x, y in ip_lan_info_dict.items():
            print(x, y)
        airties_ip = "0.0.0.0"

        for device_mac in ip_lan_info_dict:
            if airties_name == ip_lan_info_dict[device_mac]['Name']:
                airties_ip = ip_lan_info_dict[device_mac]['IP']
                print(' Airties ip:  ' + airties_ip + '\n')

        if airties_ip == "0.0.0.0":
            rf.write('    ' + 'Airties device not found in sh IP lan: Aborting \n')
            print('   Airties device not found in sh IP lan: Aborting \n')
            return "Fail"

        print(' Airties ip: ' +  airties_ip + '\n')
        print(' This is before (1) \n')
        sleep(60)
        # this doesn't work afer the commit button is pressed for some reason
        airties_cli_session =  nvg_599_dut.login_4920(airties_ip)
        print(' This is a airties session:' + str(airties_cli_session) +  '\n')
        #airties_cli_session.sendline("\n")
        #airties_cli_session.expect("#")
        airties_cli_session.sendline("wl -i wl1 radar 2")
        airties_cli_session.expect("#")
        sleep(5)
        airties_cli_session.close()
    current_5g_channel = str(nvg_599_dut.get_ui_home_network_status_value("ui_channel_5g"))
    # current_5g_channel = current_5g_channel

    # after the simulating radar detection we expect the channel to change to a non DFS channel
    print('current_5g_channel after trigger', current_5g_channel)
    if current_5g_channel in NON_DFS_CHANNELS:
        rf.write("    5G channel changed to non DFS channel:" + str(current_5g_channel) + " Pass \n" )
        print("Channel change to non DFS  Passed\n")
        print("Setting back to DFS\n")
        rf.write("    Changing back to DFS channel: 100 \n" )
        nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, 100)

        current_5g_channel = nvg_599_dut.get_ui_home_network_status_value("ui_channel_5g")
        if current_5g_channel in DFS_CHANNELS:
            result = "Current 5G:" + current_5g_channel + " is a DFS channel\n"
            rf.write("    Changing back to DFS channel: 100   Pass\n\n")
            print("Setting back to DFS passed\n\n")
        # dfs_results_file.write("Test Passed\n")
    else:
        print('test failed:Channel found:', current_5g_channel, ' expected non DFS channel')
        print('current_5g_channel', current_5g_channel)
        result = "Current 5G:" + str(current_5g_channel) + " is a DFS channel\n"
        print("result string  -dbg",result)
        rf.write("    Channel found:" +  str(current_5g_channel) +  'expected non DFS channel    Fail\n')
        test_status = "Fail"
        #results_file.write("Current 5G:" ,current_5g_channel," is not a DFS channel\n")
    return test_status
    # dfs_results_file.write(result)


def tst_speed_test(nvg_599_dut, results_file, test_ip):
    try:
        print('in tst_speed_test')
        now = datetime.today().isoformat()
        results_file.write("Test Title:tst_speed_tst Execution time:")
        results_file.write(now)
        results_file.write("\n")
        down_load_speed, up_load_speed = nvg_599_dut.speed_test_cli(test_ip)
        print('Download speed:', down_load_speed, 'Upload speed:', up_load_speed)
        speed = 'Download speed:' + down_load_speed + ' Upload speed:' + up_load_speed
        results_file.write(speed)
        results_file.write("\n")
        gw_serial_number = nvg_599_dut.serial_number
        gw_info = 'NVG 599 serial number:' + gw_serial_number
        results_file.write(gw_info)
        results_file.write("\n")
        software_version = 'Software Version:' + nvg_599_dut.software_version
        results_file.write(software_version)
        results_file.write("\n\n\n")
    except Exception as e:
        print("type error: " + str(e))
        print(traceback.format_exc())

# deprecated
def tst_ping_ip(nvg_599_dut, ping_history_file, remote_ip):
    try:
        print('in tst_ping')
        ping_history_file = open(ping_history_file, 'a+')
        now = datetime.today().isoformat()
        ping_history_file.write("Test Title:tst_ping Execution time:")
        ping_history_file.write(now)
        ping_history_file.write("\n")
        min_ping, avg_ping, max_ping ,mdev_ping = nvg_599_dut.ping_from_local_host(remote_ip)
        min_str = 'Min time: ' + min
        ping_history_file.write(min_str)

        ping_history_file.write("\n")
        avg_str = 'Avg time: ' + avg_ping
        ping_history_file.write(avg_str)

        ping_history_file.write("\n")
        max_str = 'Max time:' + max_ping
        ping_history_file.write(max_str)

        ping_history_file.write("\n")
        mdev_str = 'mdev time:' + mdev_ping
        ping_history_file.write(mdev_str)

        ping_history_file.write("\n")
        print('min: ', min_ping)
        print('avg: ', avg_ping)
        print('max: ', max_ping)
        print('mdev:', mdev_ping)
        return "Pass"
    except Exception as e:
        print("type error: " + str(e))
        print(traceback.format_exc())

def test_ping_device_name(nvg_599_dut, device_name_to_ping):
    sh_ip_lan_dict = Nvg599Class.get_rg_ip_lan_info_dict(nvg_599_dut)
    print('dict type' + str(type(sh_ip_lan_dict)))
    for key in sh_ip_lan_dict:
        # print('key----------------------------------------------' + key)
        if sh_ip_lan_dict[key]['Name'] == device_name_to_ping:
            print('found it ' + sh_ip_lan_dict[key]['Name'])
            print('IP is: ' + sh_ip_lan_dict[key]['IP'])
            # we use the  mac from the IP table to get the radio band
            # sh_wi_cl_dict = Nvg599Class.cli_sh_wi_all_clients()
            nvg_599_dut.ping_from_local_host(sh_ip_lan_dict[key]['IP'])

def test_channel_channelband_combinations(band5_channel_list, band5_bandwidth_list,ip_to_ping, result_file_name):
    nvg_599_dut = Nvg599Class()
    # ui_set_band_bandwith_channel(self, band, bandwidth, channel):
    channel_band_file = open(result_file_name, 'a')
    now = datetime.today().isoformat()
    channel_band_file.write("\" + ""Test Title:Channel/Channel band test:" + "\n")
    channel_band_file.write(now + "\n")
    software_version = nvg_599_dut.software_version
    channel_band_file.write("NVG599 Firmware:" + software_version + "\n")

    for band5_channel in range(len(band5_channel_list)):
        for band5_bandwidth in range(len(band5_bandwidth_list)):
            print(' ')
            print('band bandwidth channel: ' + str(band5_channel_list[band5_channel]) +
                  "  "  + str(band5_bandwidth_list[band5_bandwidth]))
            print(' ')
            sleep(30)
            nvg_599_dut.ui_set_band_bandwith_channel('5g', band5_bandwidth_list[band5_bandwidth],
                                                     band5_channel_list[band5_channel])
            sleep(60)
            ping_results = nvg_599_dut.ping_check(ip_to_ping)
            print('Channel:' + str(band5_channel_list[band5_channel]) + " Bandwidth:"  +
                  str(band5_bandwidth_list[band5_bandwidth]) + ' Ping result:' + str(ping_results))
        # min_ping, avg_ping, max_ping, mdev_ping = nvg_599_dut.ping_from_local_host('192.168.1.77')
        # print('min_ping:'+  ping_file = open('ping_file.txt', 'a') min_ping + ' avg_ping:' + avg_ping + ' max_ping:' + max_ping)

            channel_band_file.writelines('Channel:' + str(band5_channel_list[band5_channel]) + " Bandwidth:" +
                                         str(band5_bandwidth_list[band5_bandwidth]) + ' Ping result:' + str(ping_results) + '\n')
        #                     self.serial_number + '  min_ping:' + min_ping + '  avg_ping:' +
        #                     '  max_ping:' + max_ping + '  max dev:' + mdev_ping)
    channel_band_file.close()

# experimental
def test_comprehension(firmware_599_available):
    i = 0
    for i in range(len(firmware_599_available)):
        print('list:', str(firmware_599_available[i]))
    upper_case = [item.upper() for item in firmware_599_available]
    print('upper case:', upper_case)


##################################################################################
# use of round function example
# x = 13.73333
# x_round = round(x)
# print('x_round:',x_round)
##################################################################################
import matplotlib.pyplot as plt
import matplotlib.image as img
#
# firmware_599_available =['/home/palmer/Downloads/nvg599-9.2.2h12d9_1.1.bin',
#     '/home/palmer/Downloads/nvg599-9.2.2h12d9_1.1.bin'
#     '/home/palmer/Downloads/nvg599-9.2.2h12d10_1.1.bin',
#     '/home/palmer/Downloads/nvg599-9.2.2h12d13_1.1.bin',
#     '/home/palmer/Downloads/nvg599-9.2.2h13d1_1.1.bin'
#     '/home/palmer/Downloads/nvg599-9.2.2h13d2_1.1.bin'
#     '/home/palmer/Downloads/nvg599-9.2.2h13d3_1.1.bin'
#     '/home/palmer/Downloads/nvg599-9.2.2h13d4_1.1.bin'                         ]


def speed_test_graph_2_devices_plt():
    tablet_download_list = [27.7, 28.9, 27.6, 27.1, 30.1, 29.2, 30.6, 28.1, 29.0, 26.2]
    note8_download_list = [55.3, 54.4, 50.1, 58.2, 48.8, 41.1, 44.0, 44.0, 50.5, 51.7]
    tablet_upload_list = [2.4, 2.3, 2.4, 2.4, 2.1, 2.3, 2.1, 2.2, 2.4, 2.1]
    note8_upload_list = [2.1, 2.1, 2.0, 2.1, 2.4, 2.1, 2.01, 2.1, 2.2, 1.9]
    plt.figure("Test House Speedtest results for NVG 599", figsize=(8, 7))
    # x_label_list = ['9.2.2h12d9','9.2.2h12d10','9.2.2h12d13','9.2.2h13d1 ',
    # '9.2.2h13d2','9.2.2h13d3', '9.2.2h13d4','9.2.2h13d5','9.2.2h13d6']
    # start number,stop number, steps
    ty = arange(1.0, 11.0, 1)
    # tx =  arange(1.0, 10.0, 1)
    # plt.xticks([1,2,3,4,5,6,7,8,9])
    # plt.xticks([1,2,3,4,5,6,7,8,9],x_label_list,rotation='vertical')
    # plt.subplots(2,1,sharey=True)

    subplot(2, 1, 1)
    plt.ylabel('Speed in MBPS')
    title('NVG 599 Download Speed')
    plt.ylim([10, 70])
    plot(ty, tablet_download_list, 'g')
    plot(ty, note8_download_list, 'b')
    plt.grid()
    #
    # plt.legend(bbox_to_anchor=(1, -1.2), ncol=1)
    # plt.legend(bbox_to_anchor=(1, -1.2), ncol=1)

    subplot(2, 1, 2)
    title('NVG 599 Upload Speed')
    plt.ylabel('Speed in MBPS')
    plt.xlabel('Firmware')
    plt.xticks([1, 2, 3, 4, 5, 6, 7, 8, 9, 10], firmware_599_names, rotation='vertical')

    plt.ylim([0, 4])

    plot(ty, tablet_upload_list, 'b', label="Samsung Note8 5 G")
    plot(ty, note8_upload_list, 'g', label="Samsung Tablet 2.4 G")
    plt.legend(bbox_to_anchor=(1, -1.0), ncol=1)
    plt.grid()
    # plt.tight_layout()
    plt.subplots_adjust(left=0.1, right=0.9, bottom=0.4, top=0.9, hspace=0.4)
    plt.legend(bbox_to_anchor=(1, -1.2), ncol=1)
    # now = datetime.today().isoformat()
    # now = datetime.today().isoformat()
    # channel_band_file.write(now + ""Test Title:Channel/Channel band test:" +"\n")
    plt.savefig("/home/palmer/Downloads/speedtest:")
    show()

# note8_file.write("Speedtest results for Samsung Note 8:" + "\n")
# note8_file.write(now + "\n")
# tablet_file.write("Speedtest results for Samsung Tablet:" + "\n")
# tablet_file.write(now + "\n")
# #note8_file.write("\" + Speedtest  + "\n")
# #channel_band_file.write(now + "\n")
#
# speed_test_result_list_tablet= []
# speed_test_result_list_tablet_download= []
# speed_test_result_list_tablet_upload= []
#
# speed_test_result_list_note8 = []
# speed_test_result_list_note8_download= []
# speed_test_result_list_note8_upload= []

# pfp
# import matplotlib.pyplot as plt
# import numpy as np
# import matplotlib.pyplot as plt
from pylab import *

# used with plt graphing
def test_rg_upgrade_speedtest(nvg_599_dut, firmware_599_available, firmware_599_names):
    ''' returns tablet_download_list, tablet_upload_list, note8_download_list, note8_upload_list'''
    now = datetime.today().isoformat()
    start_time = datetime.today().strftime("%b-%d-%Y--%H:%M")
    # x = [1, 2, 3, 4, 5, 6, 7, 8]
    note8_file_name = "note8_" + start_time + '.txt'
    tablet_file_name = "tablet_" + start_time + '.txt'

    # note8_file = open('note_file.txt', 'a')
    # tablet_file = open('tablet_file.txt', 'a')
    note8_file = open(note8_file_name, 'a')
    tablet_file = open(tablet_file_name, 'a')
    tablet_file.write(start_time + "\n")
    note8_file.write(start_time + "\n")
    note8_file.close()
    tablet_file.close()
    tablet_ping_list = []
    note8_ping_list = []
    tablet_download_list = []
    tablet_upload_list = []
    note8_download_list = []
    note8_upload_list = []

    for firmware in firmware_599_available:
        print('in firmware loop ******************* firmware' + str(firmware))
        nvg_599_dut.update_rg(firmware)
        sleep(300)
        # .67 is  the tablet
        # speed_test_result_tuple_tablet = Nvg599Class.run_speed_test_from_android_termux("192.168.1.67")
        ping_results = nvg_599_dut.ping_check('192.168.1.67')
        # a,b = Nvg599Class.run_speed_test_from_android_termux("192.168.1.67")
        a,b = nvg_599_dut.execute_speed_test_from_android_termux("192.168.1.67")
        print('Tablet ---------------------------------------------firmware:' + str(firmware))
        # print("speed_test_result_tuple:", speed_test_result_tuple_tablet)
        # for a, b in speed_test_result_tuple_tablet:
        print('tablet download speed: ' + a + ' Upload speed:' + b)
        tablet_download_list.append(a)
        tablet_upload_list.append(b)
        pickle.dump(tablet_download_list, open("tabletdownloadpickle.p", "wb"))
        pickle.dump(tablet_upload_list, open("tabletuploadpickle.p", "wb"))
        tablet_file = open(tablet_file_name, 'a')
        tablet_file.write("firmware:" + str(firmware) + "Download:" + a + " Upload:" + b + "\n")
        tablet_file.close()
        sleep(30)
        # mina, avga, maxa, mdeva = nvg_599_dut.ping_from_local_host("192.168.1.67")
        # print('192.168.1.67 - min:' + str(mina) + ' max:' + str(maxa) + ' avg: ' + avga)
        # for a, b in speed_test_result_tuple_tablet:
        #     print('tablet download speed: ' + a + ' Upload speed:' + b)
        #     tablet_download_list.append(a)
        #     tablet_upload_list.append(b)
        #     tablet_file.write("firmware:" + str(firmware) + "download:" + a + "Upload:" + b + "\n")
        #     sleep(60)

        # speed_test_result_tuple_note8 = Nvg599Class.run_speed_test_from_android_termux("192.168.1.70")
        ping_results = nvg_599_dut.ping_check('192.168.1.70')
        a,b = nvg_599_dut.execute_speed_test_from_android_termux("192.168.1.70")
        # for a, b in speed_test_result_tuple_note8:
        print('Note 8 ---------------------------------------------firmware:' + str(firmware))
        print('Note 8 download speed: ' + a + ' Upload speed:' + b)
        note8_download_list.append(a)
        note8_upload_list.append(b)
        pickle.dump(note8_download_list, open("note8downloadpickle.p", "wb"))
        pickle.dump(note8_upload_list, open("note8uploadpickle.p", "wb"))

        note8_file = open(note8_file_name, 'a')
        note8_file.write("firmware:" + str(firmware) + "Download:" + a + " Upload:" + b + "\n")
        note8_file.close()
        sleep(200)

        # # post pone this for now
        # mina, avga, maxa, mdeva = nvg_599_dut.ping_from_local_host("192.168.1.70")
        # print('192.168.1.70 - min:' + str(mina) + ' max:' + str(maxa) + ' avg: ' + avga)
        #

        # for a, b in speed_test_result_tuple_note8:
        # #for a, b in speed_test_result_tuple_note8:
        #     print('Note 8 download speed: ' + a + ' Upload speed:' + b)
        #     note8_download_list.append(a)
        #     note8_upload_list.append(b)
        #     note8_file.write("firmware:" + str(firmware) + "download:" + a + "Upload:" + b + "\n")
        #     sleep(60)

    # tablet_file.close()
    # tablet_download_list
    # note8_download_list
    plt.figure("Test House Speedtest results for NVG 599 ", figsize=(8, 7))
    ty = arange(1.0, 12.0, 1)
#    ty = arange(1.0, 10.0, 1)
    subplot(2, 1, 1)
    plt.ylabel('Speed in MBPS')
    title('NVG 599 Download Speed')
    plot(ty, tablet_download_list, 'g')
    plot(ty, note8_download_list, 'b')
    plt.grid()
    #
    subplot(2, 1, 2)
    title('NVG 599 Upload Speed')
    plt.ylabel('Speed in MBPS')
    plt.xlabel('Firmware')
    plt.xticks([1, 2, 3, 4, 5, 6, 7, 8, 9, 10], firmware_599_names, rotation='vertical')

    plot(ty, tablet_upload_list, 'b', label='Tablet')
    plot(ty, note8_upload_list, 'g', label='Note8')
    plt.legend(bbox_to_anchor=(1, -1.2), ncol=1)

    plt.grid()
    plt.tight_layout()
    plt.subplots_adjust(left=0.1, right=0.9, bottom=0.2, top=0.9)
    plt.savefig("/home/palmer/Downloads/speedtest:" + start_time)
    plt.savefig("thisisit")
    show()
    return tablet_download_list, tablet_upload_list, note8_download_list, note8_upload_list

# now = datetime.today().isoformat()
# start_time = datetime.today().strftime("%b-%d-%Y--%H:%M")
# x = [1, 2, 3, 4, 5, 6, 7, 8]
# note8_file_name = "note8_" + start_time +'.txt'
# tablet_file_name = "tablet" + start_time +'.txt'
# print('fn:',note8_file_name)
# exit()


firmware_599_names = [
        '9.2.2h12d9',
        '9.2.2h12d10',
        '9.2.2h12d13',
        '9.2.2h13d1',
        '9.2.2h13d2',
        '9.2.2h13d3',
        '9.2.2h13d4',
        '9.2.2h13d5',
        '9.2.2h13d6',
        '9.2.2h13d7',
        '9.2.2h13d8'
        ]

firmware_599_available = ['/home/palmer/Downloads/nvg599-9.2.2h12d9_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h12d10_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h12d13_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h13d1_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h13d2_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h13d3_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h13d4_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h13d5_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h13d6_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h13d7_1.1.bin',
                          '/home/palmer/Downloads/nvg599-9.2.2h13d8_1.1.bin',
                          ]

# Nvg599Class.run_speed_test_from_android_termux('192.168.1.67')
# Nvg599Class.get_wifi_info_from_android_termux('192.168.1.70')


# Nvg599Class.wait_for_ssh_to_be_ready('192.168.1.70', '8022', '20', '1')
# exit()

# def pickle_test():
#     note8_upload_list = [2.1, 2.1, 2.0, 2.1, 2.4, 2.1, 2.01, 2.1, 2.2, 1.9]
#     pickle.dump(note8_upload_list, open("mypickle.p","wb"))
#     new_pickle= pickle.load( open("mypickle.p","rb"))
#     print('new_pickle',new_pickle)
# pickle_test()
# print('done')
# exit()

# return min_ping, avg_ping, max_ping, mdev_ping
#mina, avga, maxa, mdeva  =  nvg_599_dut.ping_from_local_host("192.168.1.70")
#print('min:' + str(mina) + ' max:' + str(maxa) +  ' avg: ' + avga)
#exit()
# ################  This is the one I want
# now = datetime.today().strftime("%B %d, %Y,%H:%M")
# ########################################

def tst_ping_rg_power_level(nvg_599_dut, remote_ip, number_of_pings):
# def tst_ping_rg_power_level(nvg_599_dut, remote_ip, number_of_pings):
    ping_file = open('tst_ping_with_power_change.txt', 'a')
    ping_file.writelines('tst_ping_rg_power_level' + '\n')
    now = datetime.today().strftime("%B %d, %Y,%H:%M")
    ping_file.writelines('Date:' + now + '  599 FW Ver:' + nvg_599_dut.software_version + '  Ser. No:' + nvg_599_dut.serial_number + '\n\n')
    nvg_599_dut.disable_enable_wifi_2_4g('On')
    sleep(60)
    nvg_599_dut.disable_enable_wifi_5g('On')
    sleep(60)
    nvg_599_dut.set_wifi_power_level('band2', '100')
    nvg_599_dut.set_wifi_power_level('band5', '100')
    nvg_599_dut.disable_enable_wifi_5g('Off')
    sleep(120)
    min_ping, avg_ping, max_ping, mdev_ping, sent, received, loss = nvg_599_dut.ping_from_local_host(remote_ip,number_of_pings)
    print('min:' + min_ping + ' max:' + max_ping)
    ping_file.writelines('Band:' + 'band2' + '  Ping: ' + remote_ip + '  RG Pwr: 100%' +
                     '  Sent:' + sent + '  Received:' + received + '  Percent loss:' + loss + '%\n')
    ping_file.writelines('Minimum:' + min_ping + '  Average::' + avg_ping + '  Maximum:' + max_ping + ' Max dev:' + mdev_ping + '\n')
    ping_file.writelines('\n')
    print('xx2')
    nvg_599_dut.set_wifi_power_level('band2', '50')
    sleep(120)
    min_ping, avg_ping, max_ping, mdev_ping, sent, received, loss = nvg_599_dut.ping_from_local_host(remote_ip,number_of_pings)

    ping_file.writelines('Band:' + 'band2' + '  Ping: ' + remote_ip + '  RG Pwr: 50%' +
                      '  Sent:' + sent + '  Received:' + received + '  Percent loss:' + loss + '%\n')
    ping_file.writelines('Minimum:' + min_ping + '  Average::' + avg_ping + '  Maximum:' + max_ping + ' Max dev:' + mdev_ping + '\n')
    print('xx2')
    # restore  band2
    ping_file.writelines('\n')
    # nvg_599_dut.disable_enable_wifi_2_4g('On')
    nvg_599_dut.set_wifi_power_level('band2', '100')
    sleep(120)
    # 5g section (band 5)
    # nvg_599_dut.disable_enable_wifi_5g('On')
    nvg_599_dut.disable_enable_wifi_5g('On')
    sleep(120)
    nvg_599_dut.set_wifi_power_level('band5', '100')
    sleep(120)
    nvg_599_dut.disable_enable_wifi_2_4g('Off')
    sleep(120)

    min_ping, avg_ping, max_ping, mdev_ping, sent, received, loss = nvg_599_dut.ping_from_local_host(remote_ip,number_of_pings)
    print('min:' + min_ping + ' max:' + max_ping)

    ping_file.writelines('Band:' + 'band5' + '  Ping: ' + remote_ip + '  RG Pwr: 100%' +
                     '  Sent:' + sent + '  Received:' + received + '  Percent loss:' + loss + '%\n')
    ping_file.writelines(
        'Minimum:' + min_ping + '  Average::' + avg_ping + '  Maximum:' + max_ping + ' Max dev:' + mdev_ping + '\n')
    ping_file.writelines('\n')
    print('xxx3')
    nvg_599_dut.set_wifi_power_level('band5', '50')
    sleep(120)
    min_ping, avg_ping, max_ping, mdev_ping, sent, received, loss = nvg_599_dut.ping_from_local_host(remote_ip, number_of_pings)

    ping_file.writelines('Band:' + 'band5' + '  Ping: ' + remote_ip + '  RG Pwr: 50%' +
                     '  Sent:' + sent + '  Received:' + received + '  Percent loss:' + loss + '%\n')
    ping_file.writelines('Minimum:' + min_ping + '  Average::' + avg_ping + '  Maximum:' + max_ping + ' Max dev:' + mdev_ping + '\n')
    ping_file.writelines('\n')
    print('xxx4')
    nvg_599_dut.set_wifi_power_level('band5', '100')
    # sleep(90)

    ping_file.writelines('-----------------------------------------------------------------------------------' + '\n')
    ping_file.writelines('----------------------------------------------------------------------------------' + '\n')

    ping_file.writelines('\n')
    ping_file.close()

    nvg_599_dut.disable_enable_wifi_5g('On')
    nvg_599_dut.disable_enable_wifi_2_4g('On')
    # band = 'band5'
    # percentage = 100
    nvg_599_dut.set_wifi_power_level('band2', '100')
    # band = 'band2'
    nvg_599_dut.set_wifi_power_level('band5', '100')

def tst_android_speed_test(nvg_599_dut, remote_ip):
    # this test logins to the remote android device and executes a speed test
    # def tst_ping_rg_power_level(nvg_599_dut, remote_ip, number_of_pings):
    tst_android_speed_file = open('tst_android_speed_test.txt', 'a')
    tst_android_speed_file.writelines('tst_android_speed_test: Remote IP:' + remote_ip + '\n')
    now = datetime.today().strftime("%B %d, %Y,%H:%M")
    tst_android_speed_file.writelines('Date:' + now + '  599 FW Ver:' + nvg_599_dut.software_version +
        '  Ser. No:' + nvg_599_dut.serial_number + '\n')
    nvg_599_dut.disable_enable_wifi_5g('Off')
    sleep(120)
    # down_load_speed, up_load_speed = Nvg599Class().execute_speed_test_from_android_termux(remote_ip)
    down_load_speed, up_load_speed = nvg_599_dut.execute_speed_test_from_android_termux(remote_ip)

    tst_android_speed_file.writelines(
        'Band 2 Download Speed:' + down_load_speed + 'Band 2 Upload speed:' + up_load_speed + '\n\n')

    nvg_599_dut.disable_enable_wifi_5g('On')
    sleep(120)

    # Band 5
    nvg_599_dut.disable_enable_wifi_2_4g('Off')
    sleep(120)

    down_load_speed, up_load_speed = Nvg599Class().run_speed_test_from_android_termux(remote_ip)
    tst_android_speed_file.writelines(
        'Band 5 Download Speed:' + down_load_speed + 'Band 5 Upload speed:' + up_load_speed + '\n\n')

    nvg_599_dut.disable_enable_wifi_2_4g('On')
    sleep(120)
    tst_android_speed_file.writelines('-----------------------------------------------------------------------' + '\n')
    tst_android_speed_file.writelines('-----------------------------------------------------------------------' + '\n')
    tst_android_speed_file.writelines('\n')
    tst_android_speed_file.close()

# -pfp-
# deprecated
############################## DEPRECATED   ####################################################################################################
# def tst_ping_rg_power_level_orig(nvg_599_dut, band, percentage, remote_ip, number_of_pings):
#     print('testing ping after power level changes')
#
#     if band == 'band2':
#         nvg_599_dut.disable_enable_wifi_2_4g('On')
#         nvg_599_dut.disable_enable_wifi_5g('Off')
#     else:
#         # this is band 5
#         nvg_599_dut.disable_enable_wifi_5g('On')
#         nvg_599_dut.disable_enable_wifi_2_4g('Off')
#     nvg_599_dut.set_wifi_power_level(band, percentage)
#     min_ping, avg_ping, max_ping, mdev_ping, sent, received, loss = nvg_599_dut.ping_from_local_host(remote_ip,
#                                                     number_of_pings)
#     print('min:' + min_ping + ' max:' + max_ping)
#     ping_file = open('ping_file_with_power_change_test.txt', 'a')
#     now = datetime.today().strftime("%B %d, %Y,%H:%M")
#     ping_file.writelines('Date:' + now + '  599 FW Ver:' + nvg_599_dut.software_version + '  Ser. No:' + nvg_599_dut.serial_number + '\n')
#     ping_file.writelines('Band:' + band + '  Ping: ' + remote_ip + '  RG Pwr:' + str(percentage) + '%' +
#                          '  Sent:' + sent + '  Received:' + received + '  Percent loss:' + loss + '%\n')
#     ping_file.writelines(
#         'Minimum:' + min_ping + '  Average::' + avg_ping + '  Maximum:' + max_ping + ' Max dev:' + mdev_ping + '\n')
#     ping_file.writelines('\n')
#     ping_file.close()
#     nvg_599_dut.disable_enable_wifi_5g('On')
#     nvg_599_dut.disable_enable_wifi_2_4g('On')
#     band = 'band5'
#     percentage = 100
#     nvg_599_dut.set_wifi_power_level(band, percentage)
#     band = 'band2'
#     nvg_599_dut.set_wifi_power_level(band, percentage)

def test_auto_ssid_default_tr69_values(nvg_599_dut, ssid, rf, rfa):
    default_tr69_auto_ssid_values = nvg_599_dut.get_tr69_auto_ssid(ssid)
    test_status = "Pass"
    if (default_tr69_auto_ssid_values.find('.' + ssid + '.Enable 0') != -1):
        print('Pass ssid:' + ssid + ' Default set to 0')
        rf.write('    Pass ssid:' + ssid + ' Default set to 0 \n')
    else:
        print('Fail ssid:' + ssid + ' default not set to 0')
        rf.write('    Fail ssid:' + ssid + ' default not set to 0 \n')
        test_status = "Fail"
        # return("Fail", "Default Enable not set to 0")

    if default_tr69_auto_ssid_values.find('.' + ssid + '.Status Disabled 0') != -1:
        print('Pass ssid:' + ssid + ' Default Status set to Disabled')
        rf.write('    Pass ssid:' + ssid + ' Default Status set to Disabled \n')
    else:
        print('Fail ssid:' + ssid + ' Default Status not set to Disabled')
        rf.write('    Fail ssid:' + ssid + ' Default Status not set to Disabled \n')
        test_status = "Fail"
        # return ("Fail: SSID:" + ssid + " Status not set to Disabled")

    if default_tr69_auto_ssid_values.find('.' + ssid + '.SSID TBD') != -1:
        print('Pass ssid:' + ssid + ' Default SSID set to TBD')
        rf.write('    Pass ssid:' + ssid + ' Default SSID set to TBD \n')
    else:
        print('Fail ssid:' + ssid + ' Default SSID not set to TBD')
        rf.write('    Fail ssid:' + ssid + ' Default SSID not set to TBD \n')
        test_status = "Fail"
        # return ('Fail: SSID:' + ssid + ' Default SSID not set to TBD')
    if default_tr69_auto_ssid_values.find('.' + ssid + '.X_0000C5_DefaultSSID TBD') != -1:
        print('Pass ssid:' + ssid + ' Default SSID set to TBD')
        rf.write('Pass ssid:' + ssid + ' Default SSID set to TBD \n')
    else:
        print('Fail ssid:' + ssid + ' Default SSID not set to TBD')
        rf.write('    Fail ssid:' + ssid + ' Default SSID not set to TBD \n')
        test_status = "Fail"
        # return ('Fail: SSID:' + ssid + ' Default SSID not set to TBD')
    if default_tr69_auto_ssid_values.find('.' + ssid + '.SSID TBD') != -1:
        print('Pass ssid:' + ssid + '  SSID set to TBD')
        rf.write('    Pass ssid:' + ssid + '  SSID set to TBD \n')
    else:
        print('Fail ssid:' + ssid + ' SSID not set to TBD')
        rf.write('    Fail ssid:' + ssid + ' SSID not set to TBD \n')
        test_status = "Fail"
        # return ('Fail: SSID:' + ssid + '  SSID not set to TBD')
    if (default_tr69_auto_ssid_values.find('.' + ssid + '.X_0000C5_Encryption AESEncryption') != -1):
        print('Pass ssid:' + ssid + ' Encryption set to AESEncryption')
        rf.write('    Pass ssid:' + ssid + ' Encryption set to AESEncryption \n')
    else:
        print('Fail ssid:' + ssid + ' Default Encryption not set to AESEncryption')
        rf.write('    Fail ssid:' + ssid + ' Default Encryption not set to AESEncryption \n')
        test_status = "Fail"
        # return ('Fail: SSID:' + ssid + '  Encryption not set to AESEncryption')
    if (default_tr69_auto_ssid_values.find('.' + ssid + '.X_0000C5_Authentication WPA2PSKAuthentication') != -1):
        print('Pass ssid:' + ssid + ' Authentication set to WPA2PSKAuthentication')
        rf.write('    Pass ssid:' + ssid + ' Authentication set to WPA2PSKAuthentication \n')
    else:
        print('Fail ssid:' + ssid + ' Default Authentication not set to WPA2PSKAuthentication')
        rf.write('    Fail ssid:' + ssid + ' Default Authentication not set to WPA2PSKAuthentication \n')
        test_status = "Fail"
        # return ('Fail: SSID:' + ssid + ' Default Authentication not set to WPA2PSKAuthentication')
    if (default_tr69_auto_ssid_values.find('.' + ssid + '.X_0000C5_KePasysphrase TBD') != -1):
        print('Pass ssid:' + ssid + ' Default KeyPassphrase set to TBD')
        rf.write('    Pass ssid:' + ssid + ' Default KeyPassphrase set to TBD \n')
    else:
        print('Fail ssid:' + ssid + ' Default KeyPassphrase not set to TBD')
        rf.write('    Fail ssid:' + ssid + ' Default KeyPassphrase not set to TBD \n')
        test_status = "Fail"
        #return ('Fail: SSID:' + ssid + ' Default KeyPassphrase not set to BD')
    if (default_tr69_auto_ssid_values.find('.' + ssid + '.X_0000C5_MaxClients 3') != -1):
        print('Pass ssid:' + ssid + ' Default MaxClients set to 3')
        rf.write('    Pass ssid:' + ssid + ' Default MaxClients set to 3 \n')
    else:
        print('Fail ssid:' + ssid + ' Default MaxClients not set to 3')
        rf.write('    Fail ssid:' + ssid + ' Default MaxClients not set to 3 \n')
        test_status = "Fail"
    if (default_tr69_auto_ssid_values.find('.' + ssid + '.SSIDAdvertisementEnabled 0') != -1):
        print('Pass ssid:' + ssid + ' Default SSIDAdvertisementEnabled set to 0')
        rf.write('    Pass ssid:' + ssid + ' Default SSIDAdvertisementEnabled set to 0 \n')
    else:
        print('Fail ssid:' + ssid + ' Default SSIDAdvertisementEnabled not set to 0')
        rf.write('    Fail ssid:' + ssid + ' Default SSIDAdvertisementEnabled not set to 0 \n')
        test_status = "Fail"
        # return ('Fail: SSID:' + ssid + ' Default KepPassphrase not set to BD')
    return test_status

def get_upgrade_file_from_my_win10(nvg_599_dut,win10_laptop, test_house_devices_static_info,  rf, rfa ):
    print('get_upgrade_file_from win10_laptop')
    rf.write('get_upgrade_file_from win10_laptop' + '\n')



# test_house_devices_static_info = {
#     '88:41:FC:86:64:D6': {'device_type': 'airties_4920', 'oper_sys': 'tbd', 'radio': 'abg', 'band': '2',
#                           'state': 'None',
#                           'address_type': 'None', 'port': 'None', 'ssid': 'None', 'rssi': 'None', 'ip': 'None',
#                           'device_test_name': 'airties_1_2g', 'name': 'ATT_4920_8664D4', 'location': 'master_bedroom'},



# def test_speedtest_from_android(nvg_599_dut,device, test_house_devices_static_info, test_name,  rf, rfa ):
def test_speedtest_from_android(nvg_599_dut, device_name, test_house_devices_static_info, test_name, rf, rfa):

    print('in android speed test   device name  is:' + str(device_name))
    rf.write('Test ' + test_name +  '  Device:' + device_name + '\n')
    print('Test:' + test_name + '  Device name:' + device_name + '\n')
    # rf.write('test_speedtest_from_android' + '\n')

    # if device == 'Galaxy-Note8':
    # first get the mac of the device from the dict of known test house devices
    # I had this idea that we could use a device type and then select the first availabe device of that type
    # I decided to abandon that approach and use the device name. We want to know what device we are going to use

    # for key in test_house_devices_static_info:
    #     print('key :' + str(key) + 'device: ' + str(test_house_devices_static_info[key]['device_type']))
    #
    #     if test_house_devices_static_info[key]['device_type'] == device:
    #         device_mac = key
    #         print('device mac:' + str(device_mac))
    #         break

    # nvg_599_dut.login_nvg_599_cli()
    ip_lan_info_dict = nvg_599_dut.get_rg_ip_lan_info_dict()
    # first translate the device name to a mac using the

    for x, y in ip_lan_info_dict.items():
        print(x, y)

    device_ip = "0.0.0.0"

    for device_mac in ip_lan_info_dict:
        if device_name == ip_lan_info_dict[device_mac]['Name']:
            device_ip = ip_lan_info_dict[device_mac]['IP']

    # if 'b8:d7:af:aa:27:c3' in ip_lan_info_dict:
    # if device_mac in ip_lan_info_dict:
    if device_ip != "0.0.0.0" :
        # note_8_ip = ip_lan_info_dict['b8:d7:af:aa:27:c3']['IP']
        # device_ip = ip_lan_info_dict[device_mac]['IP']
        rf.write('    Android device IP  present in cli command sh ip lan: ' + device_ip + '\n' )
        print('device is present :' + str(device_ip))
    else:
        rf.write('    Android device IP not present in cli command sh ip lan, Aborting test \n\n')
        print('Android device IP not present in cli command sh ip lan')
        return "Fail"
    nvg_599_dut.execute_speedtest_from_android_termux(device_ip, rf, rfa)
    return "Pass"


# def upgrade_rg_cli(self, tftp_server_name, install_bin_file, rf, rfa, test_name):
def test_rg_upgrade(nvg_599_dut, tftp_server_name , install_bin_file,  rf, rfa, test_name):
    #upgrade_rg_file = '/home/palmer/Downloads/nvg599-9.2.2h13d23_1.1.bin'
    rf.write('Test ' + test_name + '\n')
    test_status = nvg_599_dut.upgrade_rg_cli(tftp_server_name, install_bin_file, rf, rfa)
    if test_status == "Fail":
        rf.write('    Fail: Upgrade failed')
        # nvg_599_dut.session_cleanup()
        return "Fail"
    else:
        rf.write('    Pass: RG upgraded to:' + install_bin_file)
        sleep(300)
        return "Pass"
# we don't expect factory reset to to fail, it should always return a duration
def execute_factory_reset(nvg_599_dut,rf,rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    reset_duration = nvg_599_dut.factory_reset_rg(rf, rfa)
    if reset_duration == "Fail":
        #rf.write('    Fail: Factory reset failed')
        print('Test:' + test_name + ":" + reset_duration + '\n\n')
        rf.write('     Test:' + test_name + "  Duration:" + reset_duration + '\n\n')
        nvg_599_dut.session_cleanup()
        return "Fail"
    else:
        sleep(200)
        print('Reset duration:' + reset_duration + '\n\n')
        rf.write('     Reset duration:' + reset_duration + '\n\n')
        return "Pass"

def setup_auto_ssid_via_tr69(nvg_599_dut,ssid, max_clients,  rf, rfa, test_name):
    rf.write('Test ' + test_name + " " + ssid + '\n')
    print('Test:' + test_name + " " +  ssid + '\n')
    setup_status = nvg_599_dut.set_auto_setup_ssid_via_tr69_cli(ssid, rf, rfa)
    if setup_status == "Fail":
        rf.write('    Fail: Auto setup ssid' + str(ssid) + 'failed\n')
    print('Test:' + test_name + " " + ssid + ":" + setup_status + '\n')
    rf.write('     Test:' + test_name + " " + ssid + ":" + setup_status + '\n\n')

def verify_auto_info_not_present_in_ui(nvg_599_dut, rf, rfa, test_name):
    status_page = nvg_599_dut.get_ui_home_network_status_page_source()
    rf.write('Test ' + test_name  + '\n')
    print('Test:' + test_name  + '\n')
    # default_tr69_auto_ssid_values = nvg_599_dut.get_tr69_auto_ssid(ssid)
    verify_auto_info_status = "Pass"
    if (status_page.find('172.16') == -1):
        print('Pass: 172.16 not found in status page')
        rf.write('     172.16 not found in status page:OK\n')
    else:
        print('Fail: 172.16  found in status page')
        rf.write('     172.16  found in status page:Fail\n')
        verify_auto_info_status = "Fail"
    status_page = nvg_599_dut.get_home_network_ip_allocation_page_source()

    if (status_page.find('172.16') == -1):
        print('Pass: 172.16.x.x  not found in IP allocation page')
        rf.write('     172.16.x.x  not found in IP allocation page:OK\n')
    else:
        print('Fail: 172.16.x.x  found in IP allocation page')
        rf.write('     172.16  found in status page:Fail\n')
        verify_auto_info_status = "Fail"
        # return("Fail", "Default Enable not set to 0")
    print('Test:' + test_name + ":" + verify_auto_info_status + '\n\n')
    rf.write('    Test:' + test_name + ":" + verify_auto_info_status + '\n\n')
    return "Fail"

def verify_auto_ssid_defaults_via_tr69(nvg_599_dut, auto_ssid_number, default_ssid, default_pass_phrase,  rf, rfa, test_name):
    tr69_output = nvg_599_dut.get_tr69_parameters_for_ssid(auto_ssid_number)
    auto_ssid_defaults_status = "Pass"
    rf.write('Test ' + test_name + " " + "SSID:" + auto_ssid_number + '\n')
    print('Test:' + test_name + " " + auto_ssid_number + '\n')

    if 'WLANConfiguration.' + auto_ssid_number + '.Enable 0' in  tr69_output:
        rf.write('     Found WLANConfiguration.' + auto_ssid_number + '.Enable 0:OK\n')
        print('Found WLANConfiguration.' + str(auto_ssid_number) + '.Enable 0:OK')
    else:
        rf.write('     Found WLANConfiguration.' + str(auto_ssid_number) + '.Enable 0: Not Found\n')
        auto_ssid_defaults_status = "Fail"


    #if 'WLANConfiguration.' + auto_ssid_number + '.SSID' + default_ssid' in  tr69_output:
    #if 'WLANConfiguration.' + auto_ssid_number + '.SSID ATT4ujR48s_REPLACEME_' in  tr69_output:

    if 'WLANConfiguration.' + auto_ssid_number + '.SSID ' + default_ssid in  tr69_output:
        #rf.write('     Found SSID ATT4ujR48s_REPLACEME_:OK\n')

        rf.write('     Found ' + default_ssid + ':OK\n')
        print('Found ' + default_ssid + ':OK\n')
    else:
        rf.write('     Default ssid ' + default_ssid + ':Not found\n')
        print('Not found ' + default_ssid + ':Fail\n')

        #rf.write('     SSID ATT4ujR48s_REPLACEME_: Not Found\n')
        auto_ssid_defaults_status = "Fail"

    if "KeyPassphrase " + default_pass_phrase in  tr69_output:
        rf.write('     Found KeyPassphrase ' + default_pass_phrase + ':OK\n')
        print('Found KeyPassphrase ' + default_pass_phrase + ':OK')
    else:
        rf.write('     KeyPassphrase ' + default_pass_phrase + ': Not Found\n')
        print('Not Found KeyPassphrase ' + default_pass_phrase + ':Fail')

        auto_ssid_defaults_status = "Fail"

    if "Authentication WPA2PSKAuthentication" in tr69_output:
        rf.write('     WPA2PSKAuthentication:OK\n')
        print('Found WPA2PSKAuthentication:OK\n')
    else:
        rf.write('     WPA2PSKAuthentication: Not Found\n')
        print('Not Found WPA2PSKAuthentication:Fail\n')

        auto_ssid_defaults_status = "Fail"

    if "MaxClients 3" in  tr69_output:
        rf.write('     MaxClients 3:OK\n')
        print('Found MaxClients 3:OK')
    else:
        rf.write('     MaxClients 3: Not Found\n')
        auto_ssid_defaults_status = "Fail"
        print('Not Found MaxClients 3:Fail')

    rf.write('     Test:' + test_name + " " + auto_ssid_number + ":" + auto_ssid_defaults_status + '\n\n')
    print('Test:' + test_name + " " + auto_ssid_number + ":" + auto_ssid_defaults_status + '\n')

    return auto_ssid_defaults_status

def ping_gw_from_4920(nvg_599_dut,rf,rfa, test_name, airties_ip = 'Default'):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # we ip_of_airties = None then we get the first associated 4920 IP.
    # this is not needed here
    ip_4920 = "None"
    if airties_ip == 'Default':
        ip_lan_dict = nvg_599_dut.get_rg_ip_lan_info_dict()
        for dict_key in ip_lan_dict:
            if ("ATT_49" in ip_lan_dict[dict_key]['Name']) and (ip_lan_dict[dict_key]['State'] == "on"):
                ip_4920 = ip_lan_dict[dict_key]["IP"]
                print('using this is 4920 ip =:' + str(ip_lan_dict[dict_key]["IP"]) + '\n')

        if ip_4920 == "None":
             rf.write('    No Airties devices in IP lan: Aborting test\n')
             print('No Airties devices  found in IP lan: Test aborted\n')
             # No point in continuing
             return "Test Aborted "
    else:
        ip_4920 = airties_ip

    # airties_session = nvg_599_dut.static_login_4920(ip_4920)
    airties_session = nvg_599_dut.login_4920(ip_4920)

    airties_session.sendline('ping -c 4 192.168.1.254')
    airties_session.expect('#')
    ping_output = airties_session.before
    print('ping output' + str(ping_output))

    if "0% packet loss" in ping_output:
        print('ping from  airties:' + str(ip_4920) + 'to RG succeeded' + str(ping_output))
        rf.write('     ping from airties:' + str(ip_4920) + ' to RG 192.168.1.254 0% packet loss:OK\n\n')
    else:
        print('ping failed')
        rf.write('     ping 192.168.1.254 failed:Fail\n\n')
        test_status = "Fail"

    airties_session.sendline('exit')
    #
    # nvg_599_dut.telnet_cli_session.close
    return test_status

def ping_airties_from_rg(nvg_599_dut,rf, rfa, test_name, airties_ip = 'Default'):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # we ip_of_airties = None then we get the first associated 4920 IP.
    # this is not needed here
    ip_4920 = "None"
    if airties_ip == 'Default':
        ip_lan_dict = nvg_599_dut.get_rg_ip_lan_info_dict()
        for dict_key in ip_lan_dict:
            if ("ATT_49"  in ip_lan_dict[dict_key]['Name']) and (ip_lan_dict[dict_key]['State'] == "on"):
                ip_4920 = ip_lan_dict[dict_key]["IP"]
                print('using this is 4920 ip =:' + str(ip_lan_dict[dict_key]["IP"]))

        if ip_4920 == "None":
            rf.write('    No Airties devices in found IP lan: Test aborted\n\n')
            print('No Airties devices in found IP lan: Test aborted \n')
            # No point in continuing
            return "Test aborted"
    else:
        ip_4920 = airties_ip
    nvg_cli_session = nvg_599_dut.login_nvg_599_cli()
    # airties_session = nvg_599_dut.static_login_4920(ip_4920)
    nvg_cli_session.sendline('ping -c 4 ' + ip_4920)
    nvg_cli_session.expect("MAGIC/UNLOCKED>")
    ping_output = nvg_cli_session.before
    print('ping output pinging 4920 frm gw' + str(ping_output))

    if "0% packet loss" in ping_output:
        print('ping airties:' + str(ip_4920) + 'from RG succeeded' + str(ping_output))
        rf.write('     ping to airties:' + str(ip_4920) + ' from RG 192.168.1.254 0% packet loss:OK\n\n')
    else:
        print('ping failed')
        rf.write('     ping  to airties' + str(ip_4920) + 'from RG 192.168.1.254 failed:Fail\n\n')
        test_status = "Fail"
    nvg_cli_session.sendline('exit')
    return test_status

def verify_airties_build_versions(nvg_599_dut,rf, rfa, test_name, default_2g_fw = 'AT.922.13.3.73', default_5g_fw = 'AT.922.13.3.74'):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    nvg_599_dut.login_nvg_599_cli()
    nvg_599_dut.telnet_cli_session.sendline('!')
    nvg_599_dut.telnet_cli_session.expect('#')
    nvg_599_dut.telnet_cli_session.sendline('cat /airties/etc/buildversion')
    nvg_599_dut.telnet_cli_session.expect('#')
    fw_2g_output = nvg_599_dut.telnet_cli_session.before
    print('2g>'+ str(fw_2g_output) + '<')
    if default_2g_fw in fw_2g_output:
        rf.write('     Airties build vers. 2g matches default:' + default_2g_fw + ':OK\n')
    else:
        rf.write('     Airties build vers. 2g build mismatch:' + default_2g_fw + ':Fail\n')
        test_status = "Fail"

    nvg_599_dut.telnet_cli_session.sendline('telnet 203.0.113.2')
    nvg_599_dut.telnet_cli_session.expect('#')
    nvg_599_dut.telnet_cli_session.sendline('/airties/usr/bin/fwversion')
    nvg_599_dut.telnet_cli_session.expect('#')
    fw_5g_output = nvg_599_dut.telnet_cli_session.before
    print('5g>'+ str(fw_5g_output) + '<')
    if default_5g_fw in fw_5g_output:
        rf.write('     Airties build vers. 5g matches default:' + default_5g_fw + ':OK\n')
    else:
        rf.write('     Airties build vers. 5g build mismatch:' + default_5g_fw + ':Fail\n')
        test_status = "Fail"
    rf.write('\n')

    nvg_599_dut.telnet_cli_session.sendline('exit')
    nvg_599_dut.telnet_cli_session.expect('#')
    nvg_599_dut.telnet_cli_session.sendline('exit')
    nvg_599_dut.telnet_cli_session.expect('>')
    nvg_599_dut.telnet_cli_session.sendline('exit')
    # nvg_599_dut.telnet_cli_session.close
    return test_status

def verify_airties_hello_packet_count_increasing(nvg_599_dut,rf, rfa, test_name, airties_ip = 'Default'):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # we ip_of_airties = None then we get the first associated 4920 IP.
    # this is not needed here
    ip_4920 = "None"
    if airties_ip == 'Default':
        ip_lan_dict = nvg_599_dut.get_rg_ip_lan_info_dict()
        for dict_key in ip_lan_dict:
            if ("ATT_49"  in ip_lan_dict[dict_key]['Name']) and (ip_lan_dict[dict_key]['State'] == "on"):
                ip_4920 = ip_lan_dict[dict_key]["IP"]
                print('using this is 4920 ip =:' + str(ip_lan_dict[dict_key]["IP"]))

        if ip_4920 == "None":
             rf.write('    No Airties devices in IP lan: Fail\n\n')
             print('No Airties devices in IP lan: Fail\n\n')
             # No point in continuing
             return "Fail"
    else:
        ip_4920 = airties_ip
    # woof
    try:
        airties_session = nvg_599_dut.login_4920(ip_4920)
    except Exception as e:
        print(str(e))
        test_status = "Fail"
        rf.write('    unexpected exception:' + str(e) + 'Fail\n\n')
        return

    airties_session.sendline('cat /proc/mesh-ng-topology | grep hello')
    airties_session.expect('#')
    hello_output = airties_session.before
    print('hello output' + str(hello_output))
    hello_count_reg_ex = re.compile(r'hello:\s+(\d+)')
    mo1 = hello_count_reg_ex.search(hello_output)
    first_count = int(mo1.group(1))
    # first_count = str(mo1.group(1))

    print('count is:' + str(mo1.group(1)))
    sleep(5)

    airties_session.sendline('cat /proc/mesh-ng-topology | grep hello')
    airties_session.expect('#')
    hello_output = airties_session.before
    print('hello output' + str(hello_output))
    hello_count_reg_ex = re.compile(r'hello:\s+(\d+)')
    mo1 = hello_count_reg_ex.search(hello_output)
    # second_count = str(mo1.group(1))
    second_count = int(mo1.group(1))
    if (second_count - first_count > 0):
        print('checking hellos on airties: second hello count:' + str(second_count) + ' is greater than first hello count ' + str(first_count) + ':OK\n')
        rf.write('     Hellos on airties: second hello count: ' + str(second_count) + ' is greater than first hello count: ' + str(first_count) + ':OK\n\n')
    else:
        print(' hellos on airties: second hello count:' + str(second_count) + ' not greater than first hello count' + str(
            first_count) + ':Fail \n')
        rf.write('     Hellos on airties: second hello count: ' + str(second_count) + ' not greater than first hello count: ' + str(
            first_count) + ':Fail\n\n')
        test_status = "Fail"
    airties_session.sendline('exit')
    return test_status

def verify_google_ping_from_rg_5g(nvg_599_dut,rf,rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    nvg_599_dut.login_nvg_599_band5_cli()
    nvg_599_dut.telnet_cli_session.sendline('ping -c 4 8.8.8.8')
    nvg_599_dut.telnet_cli_session.expect('#')
    ping_output = nvg_599_dut.telnet_cli_session.before

    if "0% packet loss" in ping_output:
        print('ping succeeded'+ str(ping_output))
        rf.write('     ping google (8.8.8.8) from GW 5g 0% packet loss:OK\n\n')
    else:
        print('ping failed')
        rf.write('     ping 8.8.8.8 failed:Fail\n\n')

    nvg_599_dut.telnet_cli_session.sendline('exit')
    nvg_599_dut.telnet_cli_session.expect('>')
    #
    # nvg_599_dut.telnet_cli_session.close
    return test_status


from rgclass import nvg_info


def login_nvg_599_band5_cli(self):
    print('In login_nvg_5g_cli')
    self.telnet_cli_session = pexpect.spawn("telnet 192.168.1.254", encoding='utf-8')
    self.telnet_cli_session.expect("ogin:")
    self.telnet_cli_session.sendline('admin')
    self.telnet_cli_session.expect("ord:")
    nvg_dac = self.device_access_code
    self.telnet_cli_session.sendline(nvg_dac)
    self.telnet_cli_session.expect(">")
    self.telnet_cli_session.sendline('magic')
    self.telnet_cli_session.expect(">")
    self.telnet_cli_session.sendline('telnet 203.0.113.2')
    self.telnet_cli_session.expect('#')
    self.telnet_cli_session.sendline('export LD_LIBRARY_PATH=/lib:/airties/lib')
    self.telnet_cli_session.expect("#")
    self.telnet_cli_session.sendline('export PATH=$PATH:/airties/usr/sbin')
    print("setting library paths for remote manager")
    self.telnet_cli_session.expect("#")
    return self.telnet_cli_session

def steering_radio_names_integration_smoke(nvg_599_dut, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    nvg_5g_session = nvg_599_dut.login_nvg_599_band5_cli()
    nvg_5g_session.sendline('steer-client --command dump | grep name')
    nvg_5g_session.expect('#')
    steer_client_output = nvg_5g_session.before
    print('steer client output' + str(steer_client_output))
    steer_client_output_lines = steer_client_output.splitlines()

    names = 0
    for line in steer_client_output_lines:
        print('line is:' + str(line))
        # steer_client_regex = re.compile(r'radio_name:\s+(wl0),\s+radio_mac\s+(\w+:\w+:\w+:\w+:\w+:\w+),\s+ssid: (\w+),')
        # steer_client_regex = re.compile(r'radio_name:\s+(wl0)')
        steer_client_regex = re.compile(r'radio_name:\s+(\w+),')


        steer_client_groups = steer_client_regex.search(line)
        # steer_test = steer_client_groups.group(1)

        if steer_client_groups == None:
            # print('first group:' + steer_client_groups.group(1))
            print('not on this line \n')
            # rf.write('     did not find any steering names:Fail\n\n')
        else:
            rf.write('     found steering radio name:' + steer_client_groups.group(1) + ':OK\n')
            print('found steering name:' + str(steer_client_groups.group(1)) + '\n')
            names += 1
            if names == 2:
                rf.write('\n')
    if names < 2:
        rf.write('     did not find any steering names:Fail\n\n')
        test_status = "Fail"
        print('did not find any steering names  \n')
    return test_status

def remote_manager_smoke(nvg_599_dut, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    nvg_5g_session = nvg_599_dut.login_nvg_599_band5_cli()

    nvg_5g_session.sendline('rm-client --command dump-post-errs-queue')
    nvg_5g_session.expect('#')
    post_errs_output = nvg_5g_session.before
    print('post errs output' + str(post_errs_output))
    post_errs_regex = re.compile(r'n=\[(\d+)\]')
    post_errs_groups = post_errs_regex.search(post_errs_output)
    print(str(post_errs_groups.group(1)))
    number_of_errs = int(post_errs_groups.group(1))

    # if number_of_errs != "0":
    if number_of_errs > 1 :
        print('remote_manager_smoke fails test')
        rf.write('     Found ' + str(number_of_errs) + 'error, No errors are allowed: Fail\n\n')
        test_status = "Fail"
    else:
        print('remote_manager_smoke passes test')
        rf.write('     Found ' + str(number_of_errs) + 'error, No errors are allowed: Pass\n\n')
    return test_status


def url_att_topology_smoke(nvg_599_dut, url_to_return, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    decoded_html = nvg_599_dut.urllib_get_rg_file(url_to_return, rf, rfa)
    rg_serial_number = nvg_599_dut.serial_number
    nvg_5g_mac = nvg_info[rg_serial_number]['mac5g']
    print ('nvg_mac_5g: ' + str(nvg_5g_mac) + '\n')
    print ('decoded: ' + str(decoded_html) +  '>end_decoded \n')
    ip_5g_side = "203.0.113.2"
    #topology_regex = re.compile(r'(ownaddr=.*?ipaddr={})'.format(ip_5g_side))
    topology_regex = re.compile(r'ownaddr=(\w+:\w+:\w+:\w+:\w+:\w+).*?(ipaddr={})'.format(ip_5g_side))
    topology_text =   topology_regex.search(decoded_html)
    print('rg_mac:' + topology_text.group(1))
    rg_mac = topology_text.group(1)
    print('ip:' + topology_text.group(2))
    ip_5g = topology_text.group(2)
    if topology_text == None:
         print('topology file fails test')
         rf.write('     topology file strings:ownaddr=' + rg_mac + 'and 5g side IP:' + ip_5g + 'not found:Fail\n\n')
         test_status = "Fail"
    else:
        print(topology_text)
        rf.write('     topology file strings:ownaddr=' + rg_mac + 'and 5g side IP:' + ip_5g + ' found :Pass\n\n')
    return test_status

def url_att_route_smoke(nvg_599_dut, url_to_return, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # decoded_html = nvg_599_dut.urllib_get_rg_file("http://192.168.1.254/ATT/topology", rf, rfa)
    decoded_html = nvg_599_dut.urllib_get_rg_file(url_to_return, rf, rfa)
    rg_serial_number = nvg_599_dut.serial_number
    nvg_5g_mac = nvg_info[rg_serial_number]['mac5g']
    print('nvg_mac_5g: ' + str(nvg_5g_mac) + '\n')
    print('decoded: ' + str(decoded_html) + '>end_decoded \n')
    route_regex = re.compile(r'ownaddr=(\w+:\w+:\w+:\w+:\w+:\w+).*?(dest=\w+:\w+:\w+:\w+:\w+:\w+)')
    route_text = route_regex.search(decoded_html)
    print('rg_mac:' + route_text.group(1))
    rg_mac = route_text.group(1)
    print('ip:' + route_text.group(2))
    dest_mac = route_text.group(2)
    if route_text == None:
        print('route file fails test')
        rf.write('     route file strings:ownaddr=' + rg_mac + 'and dest mac::' + dest_mac + 'not found:Fail\n\n')
        test_status = "Fail"
    else:
        print(route_text)
        rf.write('     route file strings:ownaddr=' + rg_mac + 'and dest mac:' + dest_mac + ' found :Pass\n\n')
    return test_status

def url_att_friendly_info_smoke(nvg_599_dut, url_to_return, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # decoded_html = nvg_599_dut.urllib_get_rg_file("http://192.168.1.254/ATT/topology", rf, rfa)
    decoded_html = nvg_599_dut.urllib_get_rg_file(url_to_return, rf, rfa)
    rg_serial_number = nvg_599_dut.serial_number
    nvg_5g_mac = nvg_info[rg_serial_number]['mac5g']
    print('nvg_mac_5g: ' + str(nvg_5g_mac) + '\n')
    print('decoded: ' + str(decoded_html) + '>end_decoded \n')
    friendly_regex = re.compile(r'ownaddr=(\w+:\w+:\w+:\w+:\w+:\w+).*?(friendlyname=\w+)')
    friendly_text = friendly_regex.search(decoded_html)

    # this is wrong, calling this could cause an invalid operator execption
    # print('rg_mac:' + friendly_text.group(1))
    # rg_mac = friendly_text.group(1)
    # print('ip:' + friendly_text.group(2))
    # friendlyname = friendly_text.group(2)

    if friendly_text == None:
        print('friendly-info file fails test')
        rf.write('     friendly-info file strings:ownaddr=' + nvg_5g_mac +  'not found:Fail\n\n')
        test_status = "Fail"
    else:
        print(friendly_text)
        print('rg_mac:' + friendly_text.group(1))
        rg_mac = friendly_text.group(1)
        print('ip:' + friendly_text.group(2))
        friendlyname = friendly_text.group(2)
        rf.write('     friendly-info file strings:ownaddr=' + rg_mac + 'and friendly name:' + friendlyname + ' found :Pass\n\n')
    return test_status

def url_att_cca_smoke(nvg_599_dut, url_to_return, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # decoded_html = nvg_599_dut.urllib_get_rg_file("http://192.168.1.254/ATT/topology", rf, rfa)
    # nvg_599_dut = WebDriverWait(nvg_599_dut, 10)
    decoded_html = nvg_599_dut.urllib_get_rg_file(url_to_return, rf, rfa)
    # decoded_html = re.escape(decoded_html)
    rg_serial_number = nvg_599_dut.serial_number
    nvg_5g_mac = nvg_info[rg_serial_number]['mac5g']
    print('nvg_mac_5g: ' + str(nvg_5g_mac) + '\n')
    print('decoded: ' + str(decoded_html) + '>end_decoded \n')
    # cca5g_regex = re.compile(r'label.*?(data)')
    # status_info_reg_ex = re.compile(r'(<config\sversion.*?</config>)', re.DOTALL)
    cca5g_regex = re.compile(r'(label.*data)', re.DOTALL)

    cca5g_text = cca5g_regex.search(decoded_html)
    #print('rg_mac:' + cca5g_text.group(1))
    # cca5g = cca5g_text.group(1)
    #print('ip:' + cca5g_text.group(2))
    #friendlyname = friendly_text.group(2)
    if cca5g_text == None:
        print('cca file strings: "label" and "data" :not found:Fail')
        rf.write('     cca file strings: "label" and "data" :not found:Fail\n\n')
        test_status = "Fail"
    else:
        print('cca5g text:' + cca5g_text.group(1))
        cca5g = cca5g_text.group(1)
        print('cca file strings: "label" and "data" : found:Pass\n')
        rf.write('     cca file strings:' + cca5g + ':found:Pass\n\n')
    return test_status

def url_att_steer_smoke(nvg_599_dut, url_to_return, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # decoded_html = nvg_599_dut.urllib_get_rg_file("http://192.168.1.254/ATT/topology", rf, rfa)
    # nvg_599_dut = WebDriverWait(nvg_599_dut, 10)
    decoded_html = nvg_599_dut.urllib_get_rg_file(url_to_return, rf, rfa)
    #decoded_html = re.escape(decoded_html)
    rg_serial_number = nvg_599_dut.serial_number
    nvg_5g_mac = nvg_info[rg_serial_number]['mac5g']
    print('steer: ' + str(nvg_5g_mac) + '\n')
    print('decoded: ' + str(decoded_html) + '>end_decoded \n')
    steer_regex = re.compile(r'ownaddr=(\w+:\w+:\w+:\w+:\w+:\w+).*?(steering)')
    steer_text = steer_regex.search(decoded_html)
    # print('rg_mac:' + cca5g_text.group(1))
    # cca5g = cca5g_text.group(1)
    # print('ip:' + cca5g_text.group(2))
    # friendlyname = friendly_text.group(2)
    if steer_text == None:
        print('steer file  :not found:Fail')
        rf.write('     steer file strings :not found:Fail\n\n')
        test_status = "Fail"
    else:
        print('steer:' + steer_text.group(1))
        print('steer file strings:' + steer_text.group(1) + ' ' + steer_text.group(2) + 'found:Pass')
        rf.write('     steer file strings:' + steer_text.group(1) + ' ' + steer_text.group(2) +':found:Pass\n\n')
    return test_status

def band5_peers_cleared_after_reset(nvg_599_dut, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # decoded_html = nvg_599_dut.urllib_get_rg_file("http://192.168.1.254/ATT/topology", rf, rfa)
    # nvg_599_dut = WebDriverWait(nvg_599_dut, 10)
    # band5_cli_session  = nvg_599_dut.login_nvg_599_5g_cli()
    peer_list = nvg_599_dut.get_rg_airties_band5_wds_links(rf, rfa)
    if len(peer_list) == 0:
        print('peer list empty  :found :Pass')
        rf.write('     peer list empty  ::Pass\n\n')
    else:
        print('peer list not empty  :found ' + str(peer_list) + ':Fail')
        rf.write('     peer list not empty  :found ' + str(peer_list) + ':Fail\n\n')
        test_status = "Fail"
    print(str(peer_list))

    return test_status

def band5_peers_set_after_airties_association(nvg_599_dut, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # decoded_html = nvg_599_dut.urllib_get_rg_file("http://192.168.1.254/ATT/topology", rf, rfa)
    # nvg_599_dut = WebDriverWait(nvg_599_dut, 10)
    # band5_cli_session  = nvg_599_dut.login_nvg_599_5g_cli()
    peer_list = nvg_599_dut.get_rg_airties_band5_wds_links(rf, rfa)
    if len(peer_list) == 0:
        print('peer list empty  :found :Fail')
        rf.write('     peer list empty, expected wds mac(s)  ::Fail\n\n')
        test_status = "Fail"
    else:
        print('peer list not empty  :found ' + str(peer_list) + ':Pass')
        rf.write('     peer list not empty  :found ' + str(peer_list) + ':Pass\n\n')
    print(str(peer_list))

    return test_status

# this test should be reset airties to factory default and then rejoin RG SSID via WPS
# use one of the available 4920s (ATT_4290_866D4 or ATT_4920_C356C0
def airties_wps_connection_after_airties_factory_reset(nvg_599_dut, airties_ssid, rf, rfa, test_name):
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    print('Test:' + test_name + '\n')
    # check that airties device is in RG SSID by ping
    # login to the cli and execute sh ip lan to look for the airties devie name and get it's IP

    nvg_599_dut.wps_pair_default_airties(airties_ssid)

    # we want to pass in a remote file source  in case we are getting  files from somewhere else-- do I need a put?
    # def tftp_get_file_cli(self, remote_file_source, *source_device_list):
    # def tftp_get_file_cli(self, remote_file_source, firmware_to_get):

# rg_firmware nvg599-11.5.0h0d1_1.1.bin"
def tftp_rg_firmware_and_install(nvg_599_dut, tftp_server_name, rg_firmware, rf, rfa, test_name):
    print('Test:' + test_name + '\n')
    test_status = "Pass"
    rf.write('Test ' + test_name + '\n')
    nvg_599_dut.software_version = rg_firmware
    # If we cant get the file there is no point in continuing
    test_status = nvg_599_dut.install_rg_cli(tftp_server_name, rg_firmware, rf, rfa)
    if test_status == "Pass":
        rf.write('    Successfully loaded:' + str(rg_firmware) + ' :Pass\n')
    else:
        print('Firmware load  failed')
        rf.write('     Firmware load:' + str(rg_firmware) + ' :Failed\n\n')
        return test_status
        # test_status = "Fail"
    # make sure we can access the UI
    test_status = nvg_599_dut.check_for_system_info_page()
    if test_status == "Pass":
        print('UI check  found Pass')

        rf.write('     UI check OK)     :Pass\n\n')
    else:
        print('UI check not found  failed')
        rf.write('    UI check not found)     :Fail\n\n')
        # return test_status
    return test_status

def enable_auto_setup_ssid_via_tr69_cli(nvg_599_dut, ssid_number, max_clients, rf, rfa, test_name):
    rf.write('Test ' + str(test_name) + '\n')
    test_status = nvg_599_dut.enable_auto_setup_ssid_via_tr69_cli(ssid_number, max_clients, rf, rfa)
    if test_status == "Pass":
        rf.write('     Auto SSID:'  + str(ssid_number) + 'Successfully configured :Pass\n\n')
    else:
        print('auto ssid conf fail ')
        rf.write('     Auto SSID:'  + str(ssid_number) + 'Failed configured :Fail\n\n')
    return test_status

def local_to_remote_ping(nvg_599_dut, rf, rfa, remote_ip, test_name, number_of_pings=20):
    #  = remote_ip
    rf.write('Test ' + str(test_name) + '\n')
    test_status = "Passed"
    min_ping, avg_ping, max_ping, mdev_ping, sent, received, loss = nvg_599_dut.ping_from_local_host(remote_ip, number_of_pings=20)
    if received == "0":
        rf.write('     Local to remote ping failed)     :Fail\n')
        rf.write('     min:' + min_ping + '  avg:' + avg_ping + '  max:' + max_ping + '  mdev:' + mdev_ping  + '\n')
        rf.write('     sent:' + sent + '  received:' + received + '  loss:' + loss +  '\n\n')
        print('     sent:' + str(sent) + 'received:' + str(received) + 'loss:' + str(loss) +  '\n')
        test_status = "Fail"
    else:
        rf.write('     Local to remote ping pass     :Pass\n')
        rf.write('     min:' + min_ping + ' avg:' + avg_ping + ' max:' + max_ping + ' mdev:' + mdev_ping  + '\n')
        rf.write('     sent:' + sent + ' received:' + received + ' loss:' + loss +  '\n\n')
        print('     sent:' + str(sent) + ' received:' + str(received) + ' loss:' + str(loss) +  '\n')
    return test_status
    # test_status = nvg_599_dut.set_auto_setup_ssid_via_tr69_cli(ssid_number, max_clients, rf, rfa)
    # return min_ping, avg_ping, max_ping, mdev_ping, sent, received, loss
        # test_status = "Fail"
    # make sure we can access the UI
    #test_status = nvg_599_dut.check_for_system_info_page()
# ssid = 3
# nvg_599_dut.set_auto_setup_ssid_via_tr69_cli(ssid)
# results_file.write("Test Title: Auto SSID 3 setup: Pass")
# results_file_archive.write("Test Title: Auto SSID 3 setup: Pass")
#
# ssid = 4
# nvg_599_dut.set_auto_setup_ssid_via_tr69_cli(ssid)
# results_file.write("Test Title: Auto SSID 4 setup: Pass")
# results_file_archive.write("Test Title: Auto SSID 4 setup: Pass")
# results_file.close()
# results_file_archive.close()
# nvg_599_dut.email_test_results(results_file)
# sleep(30)
# def ui_set_wifi_password(self, security, password):

#def get_rg_ip_lan_info_dict(self):
#ip_lan_connections_dict_cli
# Galaxy-S9

def xxremote_android_speed_test(nvg_599_dut, rf, rfa, remote_device, test_name):
    #  = remote_ip
    rf.write('Test ' + str(test_name) + '\n')
    # get the ip of the device
    speed_test_ip = "192.168.1.77"
    nvg_599_dut.execute_speedtest_from_android_termux(speed_test_ip, rf, rfa)
    print('execute_speedtest_from_android_termux')
    # rfa = "future"
    # excel_cell = rfa
    # prompt = '\$\s+'
    test_status = "Passed"

def install_airties_firmware(nvg_599_dut, rf, rfa, test_name, airties_firmware, remote_device_name = "Any"):
    #  = remote_ip
    rf.write('Test ' + str(test_name) + '\n')
    # remote device name is any then get the first available airties
    # else use the one provided, if not on the loal lan then error.
    # get the ip of the device from the name
    print('install_airties_firmware')
    test_status = "Passed"
    return test_status

def guest_client_cannot_ping_rg(nvg_599_dut, rf, rfa, test_name, guest_ssid, guest_password = "default_guest_password"):
    """
    Set home ssid and password to defaults so we know what they are
    Connect to guest network
    Turn off wired connectionverify we can ping google
    verify that we cannot ping the RG and turn on wired
    connect back to main ssid
    If the home_password is "default" then the Security is set to "Default Password"
    :param nvg_599_dut:
    :param rf:
    :param rfa:
    :param test_name:
    :param guest_ssid:
    :param guest_password:
    :return:
    """
    print('in auto \n')
    rf.write('Test ' + str(test_name) + '\n')
    test_status = "Pass"
    global nvg_info
    home_ssid_conf, home_password_conf = nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa, home_ssid="default", home_password="default")
    print('**************ssid default  from conf call:' + home_ssid_conf  + '  ssid default:' + home_password_conf + '\n\n')
    sleep(60)
    cmd = 'nmcli dev wifi rescan'
    nvg_599_dut.nmcli_set_connection(cmd)

    sleep(60)

    try:
        cmd = 'nmcli device wifi connect ' + home_ssid_conf + ' password ' + home_password_conf
        print('nmcli device wifi connect to default  ssid:' + str(home_ssid_conf) + ' password:' + str(home_password_conf))
        nvg_599_dut.nmcli_set_connection(cmd)
    except subprocess.CalledProcessError as e:
        print('errror ' + e.output)
        print('failed to connect back to  SSID')
        rf.write('     Failed to connect  to default home ssid.    :Fail\n')
        return "Fail"
    # eeh4jxmh7q26
    rf.write('     Turn off the wired connections     :OK\n')
    print('Turning down wired_ssid:\n')
    nmcli_cmd = "nmcli con down Wired "
    nvg_599_dut.nmcli_set_connection(nmcli_cmd)
    sleep(120)
    cmd = 'nmcli dev wifi rescan'
    nvg_599_dut.nmcli_set_connection(cmd)
    sleep(120)
    ping_status = nvg_599_dut.ping_check('192.168.1.254')
    # Make sure we can ping the ssid
    if ping_status == "Pass":
        rf.write('      STA SSID to 192.168.1.254(RG)             :Pass\n')
        print('Ping from STA Home SSID to RG 192.168.1.254 passed:Pass\n')
    else:
        rf.write('     Ping  from STA SSID to RG IP failed     :Fail\n')
        print('Ping from STA main SSID to RG IP failed      :Fail\n')

    # config the guest network and connect to it
    guest_ssid_conf, guest_password_conf = nvg_599_dut.conf_guest_network_ssid_and_password(rf, rfa, "enable", guest_ssid, guest_password)
    print('**************guest ssid defaults after  conf call:' + guest_ssid_conf  + '  ssid default:' + guest_password_conf + '\n\n')
    sleep(60)
    cmd = 'nmcli dev wifi rescan'
    nvg_599_dut.nmcli_set_connection(cmd)
    sleep(60)

    # Connect to the guest network and verify that we can ping google.com on the wan
    try:
        cmd = 'nmcli device wifi connect ' + guest_ssid_conf  + ' password ' + guest_password_conf
        print('nmcli device wifi connect to guest ssid ' +  str(guest_ssid_conf)  + ' password ' + str(guest_password_conf))
        nvg_599_dut.nmcli_set_connection(cmd)
        sleep(60)
    except subprocess.CalledProcessError as e:
        print('errror ' + e.output)
        print('failed to connect to  default SSID')
        rf.write('     Failed to connect to  default ssid.    :Fail\n')
        return "Fail"
    # Make sure we can ping the a wan website
    ping_status = nvg_599_dut.ping_check('google.com')
    if ping_status == "Fail":
        rf.write('     Local to google.com  ping failed. Cannot continue     :Fail\n')
        try:
            cmd = 'nmcli c up Wired'
            print('nmcli wired connection up')
            nvg_599_dut.nmcli_set_connection(cmd)
        except subprocess.CalledProcessError as e:
            print('errror ' + e.output)
            print('failed to restore Wired connection')
            rf.write('     Failed to restore Wired connection.    :Fail\n')
            test_status = "Fail"
        return test_status
    else:
        rf.write('     Local to RG ping Pass.     :OK\n')

    # the last step is to ping the RG from the guest network and this should fail
    ping_status = nvg_599_dut.ping_check('192.168.1.254')
    # Make sure we can ping the ssid
    if ping_status == "Fail":
        rf.write('     STA connected to Guest SSID  ping to 192.168.1.254(RG) Failed as expected.    :Pass\n')
        print('Ping from STA Home SSID to RG 192.168.1.254 Failed  as expected: Pass\n')
    else:
        rf.write('     Ping from STA SSID to RG IP passed, expected Fail: Fail\n')
        print('Ping from STA SSID to RG IP passed, expected Fail :Fail\n')
        test_status = "Fail"

    rf.write('     Turn on the wired connections     :OK\n')
    print('Turning on wired_ssid:\n')
    nmcli_cmd = "nmcli con up Wired "
    nvg_599_dut.nmcli_set_connection(nmcli_cmd)
    sleep(120)
    cmd = 'nmcli dev wifi rescan'
    nvg_599_dut.nmcli_set_connection(cmd)
    sleep(120)
    home_ssid_conf, home_password_conf = nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa, home_ssid="default", home_password="default")
    rf.write('     \n')
    cmd = 'nmcli dev wifi rescan'
    nvg_599_dut.nmcli_set_connection(cmd)
    sleep(120)
    try:
        cmd = 'nmcli device wifi connect ' + home_ssid_conf + ' password ' + home_password_conf
        print('nmcli device wifi connect to default  ssid:' + str(home_ssid_conf) + ' password:' + str(home_password_conf) + '\n\n')
        nvg_599_dut.nmcli_set_connection(cmd)
    except subprocess.CalledProcessError as e:
        print('errror ' + e.output)
        print('failed to connect back to  SSID')
        rf.write('     Failed to connect  to default home ssid.    :Fail\n')
        return "Fail"

    sleep(300)
    return test_status
    # exit()


    # get the connection lists from nmcli
    # turn on guest network  and set the ssid and password guest_password
   # #  nvg_599_dut.conf_guest_network_ssid_and_password(rf, rfa, "on", guest_ssid , guest_password = "2222222222")
   #  nvg_599_dut.conf_guest_network_ssid_and_password(rf, rfa, "on", guest_ssid , guest_password)
   #  print('change guest password and connect to it:' + guest_ssid + ':password:' + guest_password + '\n')
   #  sleep(20)
   #  cmd = 'nmcli dev wifi rescan'
   #  nvg_599_dut.nmcli_set_connection(cmd)
   #  sleep(20)
   #  try:
   #      cmd ='nmcli device wifi connect ' + guest_ssid + ' password ' + guest_password
   #      print('nmcli device wifi connect:' + str(guest_ssid) + ' password: ' + str(guest_password))
   #      nvg_599_dut.nmcli_set_connection(cmd)
   #  except subprocess.CalledProcessError as e:
   #      print('errror ' + e.output)
   #      print('failed to connect to guest SSID')
   #      print('failed nmcli device wifi connect:' + str(guest_ssid) + ' password: ' + str(guest_password))
   #      rf.write('     Failed to connect to guest ssid.    :Fail\n')
   #      test_status = "Fail"
   #      # always want to restore the Wired connection
   #      nmcli_cmd = "nmcli con up Wired "
   #      nvg_599_dut.nmcli_set_connection(nmcli_cmd)
   #      return test_status
   #
   #  rf.write('     Turn off the wired connections     :OK\n')
   #  print('turning down wired_ssid:\n')
   #  nmcli_cmd = "nmcli con down Wired "
   #  nvg_599_dut.nmcli_set_connection(nmcli_cmd)
   #  sleep(10)
   #  ping_status = nvg_599_dut.ping_check('192.168.1.254')
   #  # Make sure we can ping the ssid
   #  if ping_status == "Fail":
   #      rf.write('     Guest SSID to 192.168.1.254  failed.    :Pass\n')
   #      print(' Ping from Guest SSID to RG 192.168.1.254   failed as expected    :Pass\n')
   #  else:
   #      rf.write('     Ping guest SSID to RG IP passed     :Fail\n')
   #      print('GPing guest SSID to RG IP passed     :Fail\n')
   #
   #  try:
   #      cmd = 'nmcli c up Wired'
   #      print('nmcli wired connection up')
   #      nvg_599_dut.nmcli_set_connection(cmd)
   #  except subprocess.CalledProcessError as e:
   #      print('errror ' + e.output)
   #      print('failed to restore Wired connection')
   #      rf.write('     Failed to restore Wired connection.    :Fail\n')
   #      test_status = "Fail"
   #
   #  #ssid_default, ssid_default_password = nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa, "default", "default")
   #
   #  try:
   #      cmd = 'nmcli device wifi connect ' + ssid_default + ' password ' + ssid_default_password
   #      print('nmcli device wifi connect back to default  ssid:' + ssid_default + ' password:' + ssid_default_password )
   #      nvg_599_dut.nmcli_set_connection(cmd)
   #  except subprocess.CalledProcessError as e:
   #      print('errror ' + e.output)
   #      print('failed to connect back to  SSID')
   #      rf.write('     Failed to connect back to default ssid.    :Fail\n')
   #      return "Fail"
   #
   #  # for  connection in active_connection_list:
   #  #     nvg_599_dut.nmcli_set_connection(connection, 'up')
   #  rf.write('     _airties_firmware     :Pass\n')
   #
   #  # restoring connection to default SSID
   #  # ssid_default, ssid_default_password = nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa, "default", "default"
   #
   #  return test_status


def enable_guest_network_and_set_passwords(self, rf, rfa, test_name, ssid_password = '1111111111', guest_ssid_password = '2222222222'):
# def config_enable_guest_network_and_set_passwords(nvg_599_dut, rf, rfa, test_name):
    print('in config_enable_guest_network_and_set_passwords \n')
    rf.write('Test ' + str(test_name) + '\n')
    pass

#def load_airties_firmware(nvg_599_dut, rf, rfa, test_name, name_of_4920_or_any, firmware_to_load):
#def set_4920_to_factory_default(self, ip_of_4920):
# #names are:  ATT_4920_C356C0  or ATT_4920_8664D4
#     def get_ip_connected_airties_by_name(self,airties_name):
#
# ddogg

def set_airties_to_factory_default(nvg_599_dut, rf, rfa, test_name, name_of_4920_or_any):
    # def get_ip_connected_airties_by_name(self,airties_name):
    # names are:  ATT_4920_C356C0  or ATT_4920_8664D4
    airties_ip = None
    ip_lan_dict = nvg_599_dut.get_rg_ip_lan_info_dict()

    ip_list_4920 = nvg_599_dut.get_ip_list_of_4920s()
    if len(ip_list_4920) == 0:
        print('no 4920 ip available, abort   \n')
        rf.write('     No airties present in lan, aborting test.    :Fail\n')
        rf.write('     Failed to connect  to changed home ssid.    :Fail\n')
        rf.write('     No airties devices detected' + '\n')
        return "Fail"


    if name_of_4920_or_any == "any":
        airties_ip = ip_list_4920[0]
        print("ip_any:" + airties_ip)
        # the else we check to se if the named airties deive is on the list
    else:


        ip_list_4920 = nvg_599_dut.get_ip_list_of_4920s()
        if len(ip_list_4920) == 0:
            print('no 4920 ip available, abort   \n')
            rf.write('     No airties present in lan, aborting test.    :Fail\n')
            rf.write('     Failed to connect  to changed home ssid.    :Fail\n')
            rf.write('     No airties devices detected' + '\n')
            return "Fail"
        else:
            pass

    nvg_599_dut.get_connected_airties_ip_from_name(name_of_4920_or_any)

    ip_list_4920 = nvg_599_dut.get_ip_list_of_4920s()
    if len(ip_list_4920) == 0:
        print('no 4920 ip available, abort   \n')
        rf.write('     No airties present in lan, aborting test.    :Fail\n')
        rf.write('     Failed to connect  to changed home ssid.    :Fail\n')
        rf.write('     No airties devices detected' + '\n')
    # static_reset_4920()
    #  ip_4920 = ip_list_4920[0]
    #wl_4920_dict = nvg_599_dut.get_4920_ssid(ip_4920)
    # rf.write('     Logging in to airtis with IP' + ip_4920 + '\n')

    pass

def test_4920_login(self, rf, rfa):
    print('in test_4920_login \n')
    nvg_599_dut.air_cli_session = pexpect.spawn("telnet  192.168.1.72", encoding='utf-8')
    nvg_599_dut.air_cli_session.sendline('\n')
    print('1')
    nvg_599_dut.air_cli_session.expect("ogin:")
    print('2')
    nvg_599_dut.air_cli_session.sendline('root')
    print('3')
    nvg_599_dut.air_cli_session.expect("#")
    print('4')
    status_output = nvg_599_dut.air_cli_session.before
    print(status_output)

# nvg_599_dut.install_airties_firmware('192.168.1.68', '/home/palmer/Downloads/AirTies_7381.bin', rf, rfa)
def load_airties_firmware(nvg_599_dut, rf, rfa, test_name, name_of_airties_or_any, firmware_to_load):
    print('in load_4920_firmware \n')
    rf.write('Test ' + str(test_name) + '\n')
    test_status = "Pass"
    ip_list_4920 = nvg_599_dut.get_ip_list_of_4920s()
    print(ip_list_4920)
    if len(ip_list_4920) == 0:
        print('no 4920 ip available, abort   \n')
        rf.write('     No airties present in lan, aborting test.    :Not run\n\n')
        return
    print('ip:' + str(ip_list_4920[0]) + '\n')

    if name_of_airties_or_any == "any":
        airties_ip = ip_list_4920[0]
        uptime_before_reload = nvg_599_dut.get_4920_uptime(airties_ip)
        #nvg_599_dut.install_airties_firmware(airties_ip, "/home/palmer/Downloads/AirTies_Air4920US-AL_FW_3.67.8.3.7623.bin", rf, rfa)
    else:
        airties_ip = str(nvg_599_dut.get_4920_ip_from_named_4920(name_of_airties_or_any))
        print('aaa_name:' + name_of_airties_or_any)
        print('aaa_ip:' + airties_ip)
        #airties_ip = ip_list_4920[0]
        uptime_before_reload = nvg_599_dut.get_4920_uptime(airties_ip)
        print('     uptime_before_reload:' + str(uptime_before_reload) + '\n')
        # print('     uptime_before_reload_reload:' + str(uptime_before_reload) + '\n')
        rf.write('     uptime_before_reload:' + str(uptime_before_reload) + '\n')
        nvg_599_dut.install_airties_firmware(airties_ip, firmware_to_load, rf, rfa)
        rf.write('     Installed firmware:' + firmware_to_load + '\n')
        print('     Installed firmware:' +  str(firmware_to_load) + '\n')
        nvg_599_dut.install_airties_firmware(airties_ip, "/home/palmer/Downloads/AirTies_Air4920US-AL_FW_2.33.1.2.2112_telnet_enabled_preinstall.bin", rf, rfa)
        rf.write('     Installed telnet enable  patch \n')
        print('     Installed telnet enable  patch \n')
        uptime_after_reload_reload = nvg_599_dut.get_4920_uptime(airties_ip)
        print('     uptime_before_reload:' + str(uptime_before_reload) + '\n')
        rf.write('     uptime_before_reload:' + str(uptime_before_reload) + '\n')
        print('     uptime_after_reload_reload:' + str(uptime_after_reload_reload) + '\n')
        rf.write('     uptime_after_reload:' + str(uptime_after_reload_reload) + '\n')

def verify_ssid_change_propagated_to_airties(nvg_599_dut, rf, rfa, test_name , ssid_parm, password_parm = "default123"):
    rf.write('Test ' + str(test_name) + '\n')
    print('in verify_ssid_change_propagated_to_airties \n')
    test_status = "Pass"
    try:
        cmd = 'nmcli c up Wired'
        print('nmcli wired connection up')
        nvg_599_dut.nmcli_set_connection(cmd)
    except subprocess.CalledProcessError as e:
        print('errror ' + e.output)
        print('failed to restore Wired connection')
        rf.write('     Failed to restore Wired connection.    :Fail\n')
        test_status = "Fail"

    home_ssid_conf, home_password_conf =  nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa, ssid_parm, password_parm)
    rf.write('     Setting ssid to:' + ssid_parm + ' setting password to:' + password_parm + '\n')
    #test_status = rf.write('Test ' + str(test_name) + '\n')
    # After we configure the ssid we have to connect to it in order to login to the Airties device
    cmd = 'nmcli dev wifi rescan'
    sleep(10)
    nvg_599_dut.nmcli_set_connection(cmd)
    sleep(30)

    try:
        cmd = 'nmcli device wifi connect ' + home_ssid_conf + ' password ' + home_password_conf
        print('nmcli device wifi connect to default  ssid:' + str(home_ssid_conf) + ' password:' + str(home_password_conf))
        nvg_599_dut.nmcli_set_connection(cmd)
    except subprocess.CalledProcessError as e:
        print('errror ' + e.output)
        print('failed to connect to changed SSID')
        rf.write('     Failed to connect  to changed home ssid.    :Fail\n')
        home_ssid_conf, home_password_conf = nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa, "default","default")
        rf.write('     Setting ssid  back to defaults:' + home_ssid_conf + 'setting password to:' + home_password_conf + '\n')
        return "Fail"
    sleep(120)

    ip_list_4920 = nvg_599_dut.get_ip_list_of_4920s()
    if len(ip_list_4920) == 0:
        print('no 4920 ip available, abort   \n')
        rf.write('     No airties present in lan, aborting test.    :Fail\n')
        rf.write('     Failed to connect  to changed home ssid.    :Fail\n')
        home_ssid_conf, home_password_conf  = nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa, "default","default")
        rf.write('     No airties devices detected' + '\n')
        rf.write('     Setting ssid  back to defaults:' + home_ssid_conf + 'setting password to:' + home_password_conf + '\n')
        return "Fail"

    ip_4920 = ip_list_4920[0]
    wl_4920_dict = nvg_599_dut.get_4920_ssid(ip_4920)
    rf.write('     Logging in to airtis with IP' + ip_4920 + '\n')
    # # mo1 = status_info_reg_ex.search(status_output)
    # airties_wl_dict['ssid'] = str(mo1.group(1))
    # airties_wl_dict['mode'] = str(mo1.group(2))
    # airties_wl_dict['rssi'] = str(mo1.group(3))
    # airties_wl_dict['noise'] = str(mo1.group(4))
    # airties_wl_dict['channel'] = str(mo1.group(5))
    # airties_wl_dict['bssid'] = str(mo1.group(6)).lower()
    # cli_session.close()

    print('airties ssid = ' + str(wl_4920_dict['ssid']) + '\n')
    airties_ssid = wl_4920_dict['ssid']
    #  if the airties sid matches what was set in the RG
    if wl_4920_dict['ssid'] == home_ssid_conf :
        rf.write('     RG ssid:' + home_ssid_conf + ' =  airties ssid:' +  airties_ssid + ' :Pass \n')
        print('tada :' + str(home_ssid_conf) + '\n')
    else:
        rf.write('     RG ssid:' + home_ssid_conf + ' does not equal  airties ssid:' + airties_ssid + ' : Fail\n')
        test_status = "Fail"
        print('Fail home ssid:' + str(home_ssid_conf) +  'does not match airties:' + str(wl_4920_dict['ssid']) + '\n')
    print('     Turning up wired connection:\n')

    try:
        cmd = 'nmcli c up Wired'
        print('nmcli wired connection up')
        nvg_599_dut.nmcli_set_connection(cmd)
    except subprocess.CalledProcessError as e:
        print('errror ' + e.output)
        print('failed to restore Wired connection')
        rf.write('     Failed to restore Wired connection.    :Fail\n')
        test_status = "Fail"

    sleep(120)
    home_ssid_def, home_password_def = nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa, "default", "default")
    rf.write('     Setting RG ssid back to default: ssid:' + home_ssid_def + 'password:' + home_password_def + '\n')
    # After we configure the ssid we have to connect to it in order to login to the Airties device
    print('     setting RG ssid back to default: ssid:' + str(home_ssid_def) + 'password:' + str(home_password_def) + '\n')

    cmd = 'nmcli dev wifi rescan'
    sleep(120)
    nvg_599_dut.nmcli_set_connection(cmd)
    sleep(60)

    try:
        cmd = 'nmcli device wifi connect ' + home_ssid_def + ' password ' + home_password_def
        print('nmcli device wifi connect to default  ssid:' + str(home_ssid_def) + ' password:' + str(home_password_def))
        nvg_599_dut.nmcli_set_connection(cmd)
    except subprocess.CalledProcessError as e:
        print('errror ' + e.output)
        print('failed to connect back to  SSID:' +  str(home_ssid_def)  + '\n')
        rf.write('     Failed to connect  to default home ssid.    :Fail\n')
        return "Fail"

    sleep(20)

    return test_status


# nvg_599_dut.get_4920_ssid("192.168.1.67")


"""The get() method takes maximum of two parameters:
key - key to be searched in the dictionary
value (optional) - Value to be returned if the key is not found. The default value is None."""
def switch(choice):
    switcher = {
       'Ayushi':'Monday',
       'Megha':'Tuesday'}
    print(switcher.get(choice,'Hi, user'))

# switch("Ayushi")
# exit()


def pyfunc(r):
    for x in range(r):
        print(' '*(r-x-1)+'*'*(2*x+1))

# ssid_number, max_clients, rf, rfa, test_name
#     def set_auto_setup_ssid_via_tr69_cli(self, ssid_number, max_clients, rf, rfa, test_name):

# def conf_auto_setup_ssid_via_tr69_cli(nvg_599_dut, auto_ssid_num, rf, rfa, test_name, max_clients = 2):
#     print('in conf_auto_setup_ssid_via_tr69_cli \n')
#     rf.write('Test ' + str(test_name) + '\n')
#     test_status = "Pass"
#     nvg_599_dut.enable_auto_setup_ssid_via_tr69_cli(auto_ssid_num, rf, rfa, test_name, max_clients)
#     pass

def connect_to_auto_ssid(nvg_599_dut, auto_ssid_num, rf, rfa, test_name, auto_allowed_ip, auto_allowed_port, max_clients = 2):
    global nvg_info
    # ssid_nums are ssid_3 ssid_4
    print('in connect_to_auto_setup_ssid \n')
    rf.write('Test ' + str(test_name) + '\n')
    test_status = "Pass"
    print('auto_allowed_ip:' + str(auto_allowed_ip) + ' auto_allowed_port:' + auto_allowed_port)
    # connect to default ssid
    # turn up the wired connection
    rf.write('     Turn up the wired connections(1)     :OK\n')
    print('Turning up wired_ssid:\n')
    nmcli_cmd = "nmcli con up Wired "
    nvg_599_dut.nmcli_set_connection(nmcli_cmd)
    sleep(10)
    default_ssid, default_ssid_password = nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa)
    print('setting default ssid:' + str(default_ssid) + ' default password:' + default_ssid_password)
    sleep(20)
    cmd = 'nmcli dev wifi rescan'
    nvg_599_dut.nmcli_set_connection(cmd)
    sleep(120)
    ping_status = nvg_599_dut.ping_check('192.168.1.254')
    # Make sure we can ping the ssid
    if ping_status == "Pass":
        rf.write('      STA SSID to 192.168.1.254       :Pass\n')
        print('Ping from STA Home SSID to RG 192.168.1.254 passed:Pass\n')
    else:
        rf.write('     Ping  from STA SSID to RG IP failed     :Aborting test\n')
        print('Ping from STA SSID to RG IP failed      :Abrt\n')
        return "Fail"

    # enable auto ssid parms via tr69
    nvg_599_dut.enable_auto_setup_ssid_via_tr69_cli(auto_ssid_num, rf, rfa, test_name, max_clients)
    print('setting basic auto ssid parms auto ssid:' + auto_ssid_num)
    # nvg_info = {"228946241148656": {'model': 'nvg599', 'device_access_code': "*<#/53#1/2", 'magic': 'kjundhkdxlxr',
    #                                 'mac2g': 'd0:39:b3:60:56:f1', 'mac5g': 'd0:39:b3:60:56:f4',
    #                                 'ssid_def_pw': 'c2cmybt25dey',
    #                                 'ssid_def': 'ATTJJ25r3A', 'auto_ssid_3': 'ZipKey-PSK', 'ssid_3_pw': 'Cirrent1',
    #                                 'ssid_4': 'ATTPOC', 'auto_ssid_4_pw': 'Ba1tshop'},

    rg_serial_number = nvg_599_dut.serial_number
    default_ssid =  nvg_info[rg_serial_number]['ssid_def']
    default_ssid_pw =  nvg_info[rg_serial_number]['ssid_def_pw']

    if auto_ssid_num == "3":
        auto_ssid = nvg_info[rg_serial_number]['auto_ssid_3']
        auto_password = nvg_info[rg_serial_number]['auto_ssid_3_pw']
    elif auto_ssid_num == "4":
        auto_ssid = nvg_info[rg_serial_number]['auto_ssid_4']
        auto_password = nvg_info[rg_serial_number]['auto_ssid_4_pw']
    else:
        print('Invalid auto ssid number:' + auto_ssid_num + 'aborting test')
        test_status = "Fail"
        return "Fail"
    print('setting auto ssid:' + str(auto_ssid) + ' auto password:' + auto_password)

    # connect to the desired auto ssid
    try:
        cmd = 'nmcli device wifi connect ' + auto_ssid + ' password ' + auto_password
        print('nmcli device wifi connect to auto ssid:' + str(auto_ssid) + ' password:' + str(auto_password))
        rf.write('      STA connected to auto_ssid:' +  str(auto_ssid) +      ':Pass\n')
        nvg_599_dut.nmcli_set_connection(cmd)
    except subprocess.CalledProcessError as e:
        print('errror ' + e.output)
        print('failed to connect back to  auto SSID')
        rf.write('     Failed to connect  to auto  SSID.    :Fail\n')
        return "Fail"

    # turn off the wired connection
    rf.write('     Turn off the wired connections(1)     :OK\n')
    print('Turning down wired_ssid:\n')
    nmcli_cmd = "nmcli con down Wired "
    nvg_599_dut.nmcli_set_connection(nmcli_cmd)
    sleep(10)

    ping_status = nvg_599_dut.ping_check('192.168.1.254')
    # Make sure we cannot ping the ssid
    if ping_status == "Pass":
        rf.write('      auto SSID STA ping to 192.168.1.254  passes but should fail     :Fail\n')
        print('auto SSID STA ping to 192.168.1.254  passes but should fail: Fail \n')
    else:
        rf.write('     auto SSID STA ping to 192.168.1.254  fails as expected     :Pass\n')
        print('auto SSID STA ping to 192.168.1.254 fails as expected      :Pass\n')

    # we have to turn on the wired connection again to modify the tr69 parms via cli
 # verify we are connected to the default ssid

    active_wifi_dict = nvg_599_dut.get_active_wifi_network_info()
    print('------------------active wifi ssid is: ' +  active_wifi_dict['ssid'] + '\n\n')

    try:
        cmd = 'nmcli device wifi connect ' + default_ssid + ' password ' + default_ssid_pw
        print('nmcli device wifi connect to default ssid:' + str(default_ssid) + ' password:' + str(default_ssid_pw))
        rf.write('      STA connected back to default ssid:' +  str(default_ssid) +      ':Pass\n')

        nvg_599_dut.nmcli_set_connection(cmd)
    except subprocess.CalledProcessError as e:
        print('errror ' + e.output)
        print('failed to connect back to  default  SSID')
        rf.write('     Failed to connect back to default  SSID.    :Fail\n')
        return "Fail"
    sleep(120)


    # turn up the wired connection
    rf.write('     Turn up the wired connections(1)     :OK\n\n')
    print('Turning up wired_ssid:\n')
    nmcli_cmd = "nmcli con up Wired "
    nvg_599_dut.nmcli_set_connection(nmcli_cmd)
    sleep(60)
    
    active_wifi_dict = nvg_599_dut.get_active_wifi_network_info()
    print('---------final active wifi ssid is ' + active_wifi_dict['ssid'] + '\n\n')
    sleep(29)

#deprecated because I am going to use the dictionary approach to first get the values and then check them
    #
    # # useful utility function
    # auto_ssid_dict = nvg_599_dut.get_tr69_auto_ssid_dict("3")
    # print("------enable------>" + str(auto_ssid_dict["enable"]))
    # print("--------channel---->" + str(auto_ssid_dict["channel"]))
    # print("--------ssid---->" + str(auto_ssid_dict["ssid"]))
    # print("--------default ssid---->" + str(auto_ssid_dict["default_ssid"]))
    # print("--------password---->" + str(auto_ssid_dict["password"]))
    # print("--------default_allowed_destinations---->" + str(auto_ssid_dict["default_allowed_destinations"]))
    # print("--------default_allowed_ports---->" + str(auto_ssid_dict["default_allowed_ports"]))


# def check_auto_defaults_via_tr69_cli(nvg_599_dut, auto_ssid_num, rf, rfa, test_name):
#     print('in check_auto_defaults_via_tr69_cli \n')
#     rf.write('Test ' + str(test_name) + '\n')
#     test_status = "Pass"
#     string_to_check  = nvg_599_dut.get_auto_setup_ssid_via_tr69_cli(auto_ssid_num)
#     print('test:' + str(string_to_check))
#     # patterns = ['dog, 'cat']
#     patterns = ['dog', 'rat']
#     string_to_check = "zzzzzzzzzzzzzzz"
#     for pattern in patterns:
#         if re.search(pattern, string_to_check) != None:
#             print('found it')
#         else:
#             print('did not find it')
#     exit()
#     if re.search(r'.*zzz', string_to_check):
#     # if re.search(r'.*34\.214\.197\.127,34\.210\.237\.47', string_to_check) is not None:
#         print('---------------default whitelist ips present: Pass\n')
#     else:
#         print('default whitelist ips present: Pass\n')
#         test_status = "Fail"



a = [1, 4, 9, 16, 25, 36, 49, 64, 81, 100]
def printeven(mylist):
    newlist = [x for x in a if x % 2 ==0]
    print(newlist)

def reverse_list(list):
    reverse_list = list[:: -1]
    if reverse_list == list:
        print("this is a palindrome")
        print(reverse_list)
    else:
        print('this is not a palindrome')
        print(reverse_list)

def print_pattern(num):
    times_to_print = 1
    number_to_print = 1
    for num in range(num):
        print(str(number_to_print ) * times_to_print)
        times_to_print += 1
        number_to_print += 1

def twolists(list_odd, list_even):
    new_list_odd = [x + 2 for x in list_odd if x % 2 == 1]
    print('odd:' + str(new_list_odd))
    new_list_even = [x + 2 for x in list_even if x % 2 == 0]
    new_list_odd.extend(new_list_even)
    print('even:' + str(new_list_even))
    print('odd:' + str(new_list_odd))
    print('final:' + str(new_list_odd))

def add_one_to_list(num):
    new_list = [x +1 for x in range(num) if x > 2]
    print(str(new_list))

def list_overlap(list1, list2):
    new_list = []
    for i in list1:
        if i in list2:
            new_list.append(i)
    print(str(new_list))

def stopif237(numlist):
    newlist = [x for x in numlist if x != 237]

numbers = [
    386, 462, 47, 418, 907, 344, 236, 375, 823, 566, 597, 978, 328, 615, 953, 345,
    399, 162, 758, 219, 918, 237, 412, 566, 826, 248, 866, 950, 626, 949, 687, 217,
    815, 67, 104, 58, 512, 24, 892, 894, 767, 553, 81, 379, 843, 831, 445, 742, 717,
    958,743, 527
    ]

list1 = [1,8,3,4,5,6]
list2 = [6,7,8,9]

set1 = {0, 2, 4, 6, 8}
set2 = {1, 2, 3, 4, 5}

def python_set(set1, set2):
    print("Union", set1 | set2)
    print("intersection", set1 & set2)
    print("difference", set1 - set2)

def myfib(num):
    previous = 0
    if num == 1:
        print(0)
        return
    if num == 2:
        print(0,1)
        return

# printeven(a)
# exit()
# python_set(set1, set2)
#  exit()

from datetime import datetime

def keys_and_values(d):
    pass
# new_keys = d.keys()
# new_values = d.values()
#     for k in d:
#         print(str(k))
# # return list(new_keys), list(new_values)

# my_dict =  {"a": 1, "b": 2, "c": 3}
# (str(keys_and_values(my_dict)))
# keys_and_values(my_dict)
# exit()


def dna_complement(dna):
    comp_dict = {
        'A': 'T',
        'T': 'A',
        'C': 'G',
        'G': 'C'
    }
    comp_list = []
    for char in dna:
        comp_list.append(comp_dict[char])
    return comp_list


def first_non_repeating_letter(str):
    char_count_dict = {}
    for char in str:
        if char not in char_count_dict.keys():
            char_count_dict[char] = 1
        else:
            # char_count_dict[char] = char_count_dict.get(key,0) + 1
            char_count_dict[char] += 1

        for char in str:
            if char_count_dict[char] == 1:
                return char

# dog_list = ['aa', 'bb','cc', 'dd']
# print(str(len(dog_list)))
# for i in range(len(dog_list)):
#     print('i:'+ str(i))
#     print(str(dog_list[i]))
# print(str('----enumerate--------'))
# for index, letters in enumerate(dog_list):
#     print('index:' + str(index) + ' letters:' + (letters))

dog_list = [1,2,3,2,2,4]

def progress_days(runs):
  previous_run = 0
  progress_days = 0
  for entry in runs:
    if previous_run == 0:
      previous_run = entry
    else:
      if entry > previous_run:
        progress_days += 1
  return progress_days

# print(str(progress_days(dog_list)))

def oddish_or_evenish(num):
  print('num:' + str(num))
  int_str = str(num)
  sum = 0
  for i in int_str:
    print(i)
    sum = sum + int(i)
  if int(sum) % 2 == 0:
    print('Evenish')
  else:
    print('Oddish')


# oddish_or_evenish(43)
# oddish_or_evenish(373)
# oddish_or_evenish(4433)
def count_overlapping(intervals, point):
  count = 0
  for interval in intervals:
    start = interval[0]
    print('start:' + str(start))
    end = interval[1]
    print('end:' + str(end))
    if (point >= start) and (point <= end):
      count = count + 1
      print('count:' + str(count))
  return count

dogs = 'aaa bbb ccc ddd'
        # ddd ccc bbb aaa
# print(str(count_overlapping([[1, 2], [2, 6], [3, 5]], 5)))
#print(str(count_overlapping([[1, 2], [5, 6], [5, 7]], 5)))
# print(str(count_overlapping([[1, 2], [5, 8], [6, 9]], 7)))


# count_overlapping([[1, 5], [2, 5], [3, 6], [4, 5], [5, 6]], 5)
# count_overlapping([[1, 5], [2, 5], [3, 6], [4, 5], [5, 6]], 6)
# count_overlapping([[1, 5], [2, 5], [3, 6], [4, 5], [5, 6]], 2)
# count_overlapping([[1, 5], [2, 5], [3, 6], [4, 5], [5, 6]], 1)
import string
def move(word):
    #word_list = word.split()
    word_list = list(word)
    # print(word_list)
    # exit()
    # abc_list = string.ascii_lowercase
    abc_string = 'abcdefghijklmnopqrstuvwxyz'
    bcd_string = 'bcdefghijklmnopqrstuvwxyzA'
    abc_list = list(abc_string)
    bcd_list = list(bcd_string)
    # bcd_list = abc_list.pop(0)
    # bcd_list.append("z")
    combined_list = zip(abc_list, bcd_list)
    dict_list = dict(combined_list)
    output_list = []
    for i in word_list:
        print(str(i))
        output_list.append(dict_list[i])
    print(output_list)
    sep = ""
    print('str:' + sep.join(output_list))
    output_list = sep.join(output_list)
    # exit()
    return output_list


def add_str_nums(num1, num2):
    if num1.isdigit():
        num1_int = int(num1)
    else:
        return -1

    if num2.isdigit():
        num2_int = int(num2)
    else:
        return -1

    print(str(num1_int + num2_int))
    return str(num1_int + num2_int)

def compare_data(*args):
    print(len(args))
    if len(args) == 0 or len(args) == 1:
        return True
    print('1')

    # exit()

    # previous_arg = args[0]
    first = True
    previous = None
    for i in args:
        if first == True:
            first = False
            # print('22 previous:' + str(previous) + ' i:' + str(i))
            previous = type(i)
        else:
            # print('33 previous:' + str(previous) + ' i:' + str(i))
            if previous != type(i):
                return False
            else:
                previous = type(i)
    return True


# add_str_nums('1','2')
# move("hello")

#
# print(str(compare_data(1,[2],(3,4,4),"cat)")))
# print(str(compare_data([2],[2])))
#
# a =1
# b = 2
# print(a == a)
# print(a == b)
#
# exit()
def cap_to_front(s):
    upper_list = []
    lower_list = []
    s_list = list(s)
    for index, char in enumerate(s_list):
        print(str(index)+ " --- " + str(char))
        if char.isupper():
            upper_list.append(char)
        else:
            lower_list.append(char)
    print("  upper:" + str(upper_list))
    print("join:" + ''.join(upper_list))
    print("lower:" + str(lower_list))
    print("joined:" + ''.join(upper_list) + ''.join(lower_list))
    return ''.join(upper_list) + ''.join(lower_list)
# print(cap_to_front("dOGo"))



import subprocess


# def connect_to_auto_ssid(nvg_599_dut, auto_ssid_num, rf, rfa, test_name, auto_allowed_ip, auto_allowed_port, max_clients = 2):


def band5_channel_change_airties_ping_recovery_timer(nvg_599_dut, rf, rfa, test_name,  start_5g_channel, end_5g_channel, airties_name):

    test_status = "Pass"
    rf.write('Test ' + test_name + ' Airties device:' + airties_name + '\n')
    print('Test:' + test_name + '\n')

    rf.write('     Set channel to inial value: ' + start_5g_channel + '    :OK\n')

    session = nvg_599_dut.session
    home_link = session.find_element_by_link_text("Device")
    home_link.click()
    current_5g_channel = nvg_599_dut.get_ui_home_network_status_value("ui_channel_5g")
    # var_type = type(current_5g_channel)
    # print('type is:' + str(var_type))
    if current_5g_channel != start_5g_channel:
        nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, start_5g_channel)
        print('setting channel to :' + start_5g_channel)
        print('sleeping five minute to establish initial conditions')
        sleep(300)
    else:
        print('current channel:' + current_5g_channel + ' equals start_channel: ' + start_5g_channel)

    rf.write('     Set channel to inial value:' + start_5g_channel + '    :OK\n')

    # # This is where we change channels
    # print(' changing channels now Airties:-------------------------------------------------------\n')
    # print(' changing  Airties channel now: ' + airties_name + ' to channel ' + str(end_5g_channel) +  '\n')

    # nvg_599_dut.login_nvg_599_cli()
    ip_lan_info_dict = nvg_599_dut.get_rg_ip_lan_info_dict()
    rf.write('     Change channel to:' + end_5g_channel + '    :OK\n')

    # first translate the device name to a mac using the
    for x, y in ip_lan_info_dict.items():
        print(x, y)
    airties_ip = "0.0.0.0"
    ip_to_ping = "0.0.0.0"

    for device_mac in ip_lan_info_dict:
        if airties_name == ip_lan_info_dict[device_mac]['Name']:
            airties_ip = ip_lan_info_dict[device_mac]['IP']
            ip_to_ping = ip_lan_info_dict[device_mac]['IP']

            print(' Airties ip:  ' + airties_ip + '\n')

    if airties_ip == "0.0.0.0":
        rf.write('    ' + 'Airties device not found in sh IP lan: Aborting \n')
        print('   Airties device not found in sh IP lan: Aborting \n')
        return "Fail"
    rf.write('     Airties:' + airties_name + ' has IP:' +  str(ip_to_ping)  +  ':OK\n')

    print(' Airties ip: ' + airties_ip + '\n')
    print(' Initial band5 channel:' + start_5g_channel + ' set, now changing to  channel:' + end_5g_channel + '\n')

    # nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, end_5g_channel)

    ping_time_regex = re.compile(r'\[(\d+).+\]', re.DOTALL)
    """ We go from the start state to the inital_ping_fail state"""
    state = "start"
    ping_time_last_success = 0
    ping_time_after_failure = 0
    while True:
        if state == "start":
            n = 0
            while True:
                try:
                    # p = subprocess.check_output(['ping','192.168.1.71','-c','1',"-W","4","-D"], timeout=5).decode("utf-8")
                    p = subprocess.check_output(['ping', ip_to_ping,'-c','1',"-W","4","-D"], timeout=5).decode("utf-8")
                    print('ping count:' + str(n) + '\n')
                    n += 1
                    print(p)
                    ping_time = ping_time_regex.search(p)
                    ping_time_last_success = ping_time.group(1)
                    print('last ping_timestamp before failures:' + ping_time.group(1) + 'return_code:')
                    sleep(1)
                    if n == 1 :
                        nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, end_5g_channel)
                    sleep(1)

                except subprocess.CalledProcessError as e:
                    state = "initial_ping_fail"
                    return_code = e.returncode
                    print('return_code:' + str(state))
                    break

        elif state == "initial_ping_fail":
            n = 0
            while True:
                try:
                   #  p = subprocess.check_output(['ping','192.168.1.71','-c','1',"-W","4","-D"], timeout=5).decode("utf-8")
                    p = subprocess.check_output(['ping', ip_to_ping,'-c','1',"-W","4","-D"], timeout=5).decode("utf-8")
                    ping_time = ping_time_regex.search(p)
                    ping_time_after_failure = ping_time.group(1)
                    print('ping_timestamp:after failures: ' + ping_time.group(1) + 'return_code:')
                    n += 1
                    if n == 2:
                        state = "done"
                        sleep(1)
                        break

                except subprocess.CalledProcessError as e:
                    state = "initial_ping_fail"
                    return_code = e.returncode
                    print('state:' + state + " continuing...")
                    sleep(1)
        elif state == "done":
            # we are done and have the times needed
            break

    print("after fail:" + str(ping_time_after_failure) + " last success:" + str(ping_time_last_success))

    duration = int(ping_time_after_failure) - int(ping_time_last_success)

    rf.write('     start time:' + str(ping_time_last_success) + ' End time:' +  str(ping_time_after_failure)  +  ':OK\n')
    rf.write('     Duration:'  + str(duration) + ':OK\n\n')

    print("duration:" + str(duration))
    return duration


# ping_url_regex = re.compile(r'.*\d',re.DOTALL)
#ping_url_regex = re.compile(r'\(\d{1,3}\.', re.DOTALL)
# p = "www.google.com123"
# ping_url_regex = re.compile(r'(\d{1,3})', re.DOTALL)
# #ping_url_regex = re.compile(r'\[(\d+).+\]', re.DOTALL)
# url_ip_group = ping_url_regex.search(p)
# url_ip = url_ip_group.group(1)
# print('url ip:' + str(url_ip))
# exit()
#
# try:
#     # p = subprocess.check_output(['ping', url_to_ping, '-c', '1', "-W", "1", "-D"], timeout=5).decode("utf-8")
#     # output = subprocess.check_output(cmd, shell=True                  print('ping count:' + str(n) + '\n')
#     print('---------------------------------')
#     #print(p)
#     print('---------------------------------')
#
#
#     p = "www.google.com"
#
#     # p = "PING www.google.com (172.217.12.36) 56(84) bytes of data.[1574452795.248952] 64 bytes from dfw28s04-in-f4.1e100.net (172.217.12.36): icmp_seq=1 ttl=52 time=13.1 ms"
#
#     url_ip_group = ping_url_regex.search(p)
#     url_ip = url_ip_group.group(1)
#     print('url ip:' + str(url_ip))
# except subprocess.CalledProcessError as e:
#     print('url ip  error exception:')

# The -c means that the ping will stop afer 1 package is replied
# patches

def chatroom_status(users):
    if len(users) == 0:
        print("no users")
        return("no one online")
    elif len(users) == 1 :
        print("1 users")
        return(users[0] + " online")
    elif len(users) == 2:
        return (users[0] + " and " + users[1] + " online")
    elif len(users) > 2 :
        numusers = len(users) - 2
        mylist = []
        #my_list.append(users[0])
        my_string = users[0] + ", " + users[1] + " and " + str(numusers) +  " more online"
        # my_string = str(users[0])
        print(my_string)
        # return (users[0] + " , " + users[1]) + "and " + numusers +   " online"
        return my_string


# chatroom_status(["pap_ier44", "townieBOY", "panda321", "motor_bike5", "sandwichmaker833", "violinist91"])

def unique_sort(lst):
    list_set = set(lst)
    list_list = list(list_set)

    # list_list = list_list.sort()
    print(type(list_list))
    list_list.sort()
    return list_list

    # exit()

    # new_list = list(list_set)
    # return new_list.sort()

# exit()
# unique_sort([1, 4, 4, 4, 4, 4, 3, 2, 1, 2])
# exit()

def unique_abbrev(abbs, words):
    words_dict = {}
    for word in words:
        words_dict[word] = 0

    for k in words_dict:
        for ab in abbs:
            if k.startswith(ab):
                words_dict[k] = words_dict[k] + 1

    for k in words_dict:
        if words_dict[k] > 1 :
            return False
    return True

# print(unique_abbrev(["ho", "h", "ha"], ["house", "hope", "happy"]))
# print(unique_abbrev(["b", "c", "ch"], ["broth", "chap", "cardigan"]))
# print(unique_abbrev(["to", "too", "t"], ["topology", "took", "torrent"]))
# print(unique_abbrev(["bi", "ba", "bat"], ["big", "bard", "battery"]))

# exit()

def lines_are_parallel(l1, l2):
    if l1[1] == l2[1]:
        # print("a\n")
        return True
    else:
        new_list = [ -int(i) for i in l2]
        if l1[1] == new_list[1]:
            # print("b\n")
            #print(str(new_list))
            return True
        else:
            return False


# Test.assert_equals(lines_are_parallel([1,2,3], [1,2,4]), True, "Given example 1.")
# Test.assert_equals(lines_are_parallel([2,4,1], [4,2,1]), False, "Given example 2.")
# Test.assert_equals(lines_are_parallel([0,1,5], [0,1,5]), True, "Given example 3.")
# Test.assert_equals(lines_are_parallel([2,5,0], [20,50,10]), True)
# Test.assert_equals(lines_are_parallel([2,5,0], [-200,-500,10]), True)
# Test.assert_equals(lines_are_parallel([400000,1,0], [400000,2,0]), False)
# Test.assert_equals(lines_are_parallel([800,20,0], [40,20,0]), False)
# Test.assert_equals(lines_are_parallel([400000,1,0], [800000,2,100000]), True)
# Test.assert_equals(lines_are_parallel([-5,7,100000], [5,-7,-200000]), True)

# print(lines_are_parallel([1,2,3], [1,2,4]))
# print(lines_are_parallel([2,4,1], [4,2,1]))
# print(lines_are_parallel([0,1,5], [0,1,5]))
# print(lines_are_parallel([2,5,0], [20,50,10]))
# print(lines_are_parallel([2,5,0], [-200,-500,10]))
# print(lines_are_parallel([400000,1,0], [400000,2,0]))
# print(lines_are_parallel([800,20,0], [40,20,0]))
# print(lines_are_parallel([400000,1,0], [800000,2,100000]))
# print(lines_are_parallel([-5,7,100000], [5,-7,-200000]))
#
# exit()
# def alphabet_soup(txt):
#     new_list = [x for x in txt]
#     new_list.sort()
#     # super_new_list = new_list.join("")
#     super_new_str = "".join(new_list)
#
#
#     # print(super_new_list)
#     return super_new_str
#
# alphabet_soup("gof")
#
# exit()

def factor_chain(lst):
    for lst_item in lst:
        if len(lst) == 1 :
            return True
        if len(lst) == 2 :
            previous = lst[lst_item]
            if lst[i + 1] % previous:
                return False
            else:
                return True

    pass

# dog_list1 = ["a","b","c"]
#
# for i in range(5):
#      print(i)
# exit()

def list_of_multiples (num, length):
    # adjust length so length list can be used as multiples
    comp_test = [x+1 for x in range(length)]
    my_list = []
    for i in comp_test:
        my_list.append(i*num)
    return my_list


# print(list_of_multiples(10,5))

#my_list = [1,2,3]

# comp_test = [x+1 for x in range(5)]
#
# for i in comp_test:
#      print(i)
# exit()



#start
#import copy
from copy import deepcopy
def objectToArray (obj):
    my_list = []
    my_list_entry = []
    for key,value in obj.items():
        #print('key:' + str(key) + ' ')
        #print('value:' + str(value))
        my_list_entry.append(deepcopy(key))
        my_list_entry.append(deepcopy(value))
        print('my list entry:' + str(my_list_entry))
        my_list.append(deepcopy(my_list_entry))
        print('my list:' + str(my_list))
        my_list_entry.clear()
    return my_list

#my_dict = {'D':1, 'B':2, 'C':3}
#print('\n')
#print('final:' + str(objectToArray(my_dict)))
#exit()

# Test.assert_equals(no_duplicate_letters("Easy does it."), True)
# Test.assert_equals(no_duplicate_letters("So far, so good."), False)
# Test.assert_equals(no_duplicate_letters("Better late than never."), False)
# Test.assert_equals(no_duplicate_letters("Beat around the bush."), True)
# Test.assert_equals(no_duplicate_letters("Give them the benefit of the doubt."), False)
# Test.assert_equals(no_duplicate_letters("Your guess is as good as mine."), False)
# Test.assert_equals(no_duplicate_letters("Make a long story short."), True)
# Test.assert_equals(no_duplicate_letters("Go back to the drawing board."), True)
# Test.assert_equals(no_duplicate_letters("Wrap your head around something."), True)
# Test.assert_equals(no_duplicate_letters("Get your act together."), False)
# Test.assert_equals(no_duplicate_letters("To make matters worse."), False)
# Test.assert_equals(no_duplicate_letters("No pain, no gain."), True)
# Test.assert_equals(no_duplicate_letters("We'll cross that bridge when we come to it."), False)
# Test.assert_equals(no_duplicate_letters("Call it a day."), False)
# Test.assert_equals(no_duplicate_letters("It's not rocket science."), False)
# Test.assert_equals(no_duplicate_letters("A blessing in disguise."), False)
# Test.assert_equals(no_duplicate_letters("Get out of hand."), True)
# Test.assert_equals(no_duplicate_letters("A dime a dozen."), True)
# Test.assert_equals(no_duplicate_letters("Time flies when you're having fun."), True)
# Test.assert_equals(no_duplicate_letters("The best of both worlds."), True)
# Test.assert_equals(no_duplicate_letters("Speak of the devil."), True)
# Test.assert_equals(no_duplicate_letters("You can say that again."), False)


def no_duplicate_letters(phrase):
  # phrase = re.sub(r'[^a-zA-Z ]','',phrase)
  return_status = True
  my_word_list = phrase.split(" ")
  for word in my_word_list:
    # get rid of any non apha chars
    word = re.sub(r'[^a-zA-Z]', '', word)

    word_list = list(word)
    # my_letter_list = word.split()
    print(str(word) + ' ')
    if len(word_list) != len(set(word_list)):
      return_status = False
  return return_status


# print(str(no_duplicate_letters("easy does it")))



#my_list = [1,2,3,4,5]
#my_set_from_list = len(set(my_list))
#print('set_from_list:' + str(my_set_from_list))

# my_set = (1,2,3,4,5)
# print(len(my_set))
def digital_vowel_ban(n, ban):
    number_dict = {'1':'one','2':'two','3':'three', '4':'four', '5':'five', '6':'six','7': 'seven', '8':'eight', '9':'nine', '0':'zero'}
    number_list = list(str(n))
    number_list_copy = deepcopy(number_list)
    print('number_list:' + str(number_list))
    exit()
    for num_char in number_list:
        english_num_word = number_dict[num_char]
        english_num_word_set = set(english_num_word)
        if ban in english_num_word_set:

            pass
        # number_in_english =ban number_dict[number]
        # number_in_eglish_list = list(number_in_english.split(""))
        # number_in_eglish_set = set(number_in_english.split(""))

digital_vowel_ban(143,"o")
exit()
# Test.assert_equals(digital_vowel_ban(143, "o"), 3, "Example #1")
# Test.assert_equals(digital_vowel_ban(14266330, "e"), 4266, "Example #2")
# Test.assert_equals(digital_vowel_ban(4020, "u"), 20, "Example #3")
# Test.assert_equals(digital_vowel_ban(586, "i"), "Banned Number", "Example #4")
# Test.assert_equals(digital_vowel_ban(123456789, "i"), 12347)
# Test.assert_equals(digital_vowel_ban(20442, "o"), "Banned Number")
# Test.assert_equals(digital_vowel_ban(1100, "u"), 1100)
# Test.assert_equals(digital_vowel_ban(1993, "e"), "Banned Number")
# Test.assert_equals(digital_vowel_ban(90160350102, "e"), 62)
# Test.assert_equals(digital_vowel_ban(79284426, "o"), 7986)
# Test.assert_equals(digital_vowel_ban(123456789, "a"), 123456789, "triviAl test")



#exit()



def num_of_sublists(lst):
    # print('type:' + str(type(lst)) )
    for a in lst:
        # if a != isinstance(type(a),list):
        # print('a:', a)
        if  isinstance(a, list):
            print('type a:' + str(type(a)))
            return(len(lst))
            #print('len:' + str(len(lst)))
        else:
            return 0

    #number_of_lists = len(lst)
    #print('len:' + str(len(lst)))


# Test.assert_equals(num_of_sublists([[1,2,3], [1,2,3], [1,2,3]]), 3)
# Test.assert_equals(num_of_sublists([[1,2,3]]), 1)
# Test.assert_equals(num_of_sublists([1,2,3]), 0)
# Test.assert_equals(num_of_sublists([[1,2,3], [ 1,2,3 ], [ 1,2,3 ], [1,2,3]]), 4)
# print(num_of_sublists([1,2,3]))
# exit()
# print(num_of_sublists([[1,2,3], [1,2,3], [1,2,3]]))
# print(num_of_sublists([[1,2,3]]))
# print(num_of_sublists([1,2,3]))
# print(num_of_sublists([[1,2,3], [ 1,2,3 ], [ 1,2,3 ], [1,2,3]]))


exit()

with open('results_file.txt', mode = 'w', encoding = 'utf-8') as rf, \
    open('resultsa_file.txt', mode='w', encoding='utf-8') as rfa :
    now = datetime.today().strftime("%B %d, %Y,%H:%M")
    send_email = 1


    nvg_599_dut = Nvg599Class()
    trigger_dfs_channel_change(nvg_599_dut, rf, rfa, 'trigger_dfs_channel_change_from_airties', "ATT_4920_8664D4")

    # nvg_599_dut.login_4920("192.168.1.73")
    exit()
    # execute_factory_reset(nvg_599_dut, rf, rfa, 'execute_factory_reset')
    # exit()
    #def band5_channel_change_airties_ping_recovery_timer(nvg_599_dut, rf, rfa, test_name, start_5g_channel,
                                                         # end_5g_channel, airties_name)
    #band5_channel_change_ping_recovery_timer(nvg_599_dut, rf, rfa, "100", "149", "192.168.1.71")
    # dfs_channel_change(nvg_599_dut, rf, rfa, "dfs_channel_change", "100", "149","ATT_4920_8664D4")
    # nvg_599_dut.setup_tr69_url()
    # nvg_599_dut.login_eco()
    # steering_radio_names_integration_test(nvg_599_dut,rf,rfa,"steering_radio_names_integration_test")
    # guest_client_cannot_ping_rg(nvg_599_dut, rf, rfa, "guest_client_cannot_ping_rg", "default", "default123")
    # def url_att_cca5g_smoke(nvg_599_dut, url_to_return, rf, rfa, test_name):
   #  def remote_manager_smoke(nvg_599_dut, rf, rfa, test_name):

    #
    # check_auto_defaults_via_tr69_cli(nvg_599_dut, '3', rf, rfa,  'check_auto_defaults_via_tr69_cli')
    # nvg_599_dut.get_auto_setup_ssid_via_tr69_cli('3')
    # nvg_599_dut.rg_setup_without_factory_reset(rf, rfa)
    # exit()
    # home_ssid_conf, home_password_conf = nvg_599_dut.conf_home_network_ssid_and_password(rf, rfa, home_ssid="default", home_password="default")
    # def conf_auto_setup_ssid_via_tr69_cli(nvg_599_dut, auto_ssid_num, rf, rfa, test_name, max_clients=2):
    # (nvg_599_dut, '3', rf, rfa, 'set_auto_setup_ssid_via_tr69_cli')
    # (nvg_599_dut, "4", rf, rfa, 'set_auto_setup_ssid_via_tr69_cli')
    #status = nvg_599_dut.set_auto_setup_ssid_via_tr69_cli(nvg_599_dut, "4", rf, rfa, 'nvg_599_dut.set_auto_setup_ssid_via_tr69_cli')    # nvg_599_dut.login_eco()
    # exit()
    # ping_airties_from_rg(nvg_599_dut, rf, rfa, "ping_airties_from_RG")
    # exit()
    #nvg_599_dut.rg_setup_without_factory_reset(rf, rfa)
    # exit()
    # nvg_599_dut.tftp_get_file_cli("LP-PPALMER" , "AirTies_Air4920US-AL_FW_xxxxxxxxxxxx.bin", rf, rfa)
    rf.write('RG Test run Firmware: nvg599-11.6.0h0d1_1.1.bin  Date:' + now +  '\n\n')
    rfa.write(now + '\n')
    sleep(1)

    #load_airties_firmware(nvg_599_dut, rf, rfa, "load_airties_firmware AirTies_Air4920US-AL_FW_3.67.8.3.7623.bin", "ATT_4920_8664D4", "/home/palmer/Downloads/AirTies_Air4920US-AL_FW_3.67.8.3.7623.bin")
    #load_airties_firmware(nvg_599_dut, rf, rfa, "load_airties_firmware AirTies_Air4920US-AL_FW_3.67.8.3.7623.bin", "ATT_4920_8664D4", "/home/palmer/Downloads/AirTies_Air4920US-AL_FW_2.33.1.2.2112_telnet_enabled_preinstall.bin")

    # load_airties_firmware(nvg_599_dut, rf, rfa, "load_airties_firmware", "any", "/home/palmer/Downloads/AirTies_Air4920US-AL_FW_3.67.8.3.7623.bin")

    # execute_factory_reset(nvg_599_dut, rf, rfa, 'execute_factory_reset')

    ##  band5_peers_set_after_airties_association(nvg_599_dut, rf, rfa, "band5_peers_set_after_airties_associationn")athee
    # verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '3', 'ZipKey-PSK', 'Cirrent1', rf, rfa, "verify_auto_ssid_defaults_via_tr69")
    # ##############verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '4', 'ATTPOC', 'Ba1tshop', rf, rfa, "verify_auto_ssid_defaults_via_tr69")
    # #################depracated use enable conf_auto_setup_ssid_via_tr69_cli(nvg_599_dut, '3', rf, rfa, 'conf_auto_setup_ssid_via_tr69_cli')
    # def conf_auto_ssid_allowed_ip_and_allowed_port_via_tr69_cli(self, ssid_number, auto_allowed_ip, auto_allowed_port, rf, rfa):

    # tftp_rg_firmware_and_install(nvg_599_dut, "LP-PPALMER", "nvg599-11.6.0h0d1_1.1.bin", rf, rfa,"tftp_rg_firmware_and_install")

    band5_channel_change_airties_ping_recovery_timer(nvg_599_dut, rf, rfa, '5g airties channel change Non DFS to DFS ping recovery time', "149", "100", "ATT_4920_8664D4")
    band5_channel_change_airties_ping_recovery_timer(nvg_599_dut, rf, rfa, '5g airties channel change Non DFS to Non DFS ping recovery time', "149", "48", "ATT_4920_8664D4")
    # exit()
    band5_channel_change_airties_ping_recovery_timer(nvg_599_dut, rf, rfa, '5g airties channel change  DFS to Non DFS ping recovery time', "100", "149", "ATT_4920_8664D4")
    band5_channel_change_airties_ping_recovery_timer(nvg_599_dut, rf, rfa, '5g airties channel change  DFS to  DFS ping recovery time', "100", "132", "ATT_4920_8664D4")
    # exit()
    remote_manager_smoke(nvg_599_dut, rf, rfa, "remote_manager_smoke")
    url_att_cca_smoke(nvg_599_dut, 'ht'
                                   'tp://192.168.1.254/ATT/cca2G', rf, rfa, "url_att_cca5g_smoke")
    url_att_cca_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/cca5G', rf, rfa, "url_att_cca2g_smoke")

    connect_to_auto_ssid(nvg_599_dut, "3", rf, rfa, "connect_to_auto_ssid", "22.33.44.55", "77")
    test_speedtest_from_android(nvg_599_dut, 'Galaxy-Note8', test_house_devices_static_info, 'test_speedtest_from_android ', rf, rfa)
    steering_radio_names_integration_smoke(nvg_599_dut, rf, rfa, 'steering_radio_names_integration_test')
    url_att_steer_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/steer', rf, rfa, 'url_att_steer_smoke')
    url_att_friendly_info_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/friendly-info', rf, rfa, 'url_att_friendly_info_smoke')
    guest_client_cannot_ping_rg(nvg_599_dut, rf, rfa, "guest_client_cannot_ping_rg", "default", "default123")
    verify_airties_hello_packet_count_increasing(nvg_599_dut, rf, rfa, "verify_airties_hello_packet_count_increasing")
    ping_gw_from_4920(nvg_599_dut, rf, rfa, "ping_gw_from_4920")
    ping_airties_from_rg(nvg_599_dut, rf, rfa, "ping_airties_from_RG")
    verify_airties_build_versions(nvg_599_dut, rf, rfa, 'verify_airties_build_versions')
    url_att_route_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/route', rf, rfa, 'url_att_route_smoke')
    url_att_topology_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/topology', rf, rfa, 'url_att_topology_smoke')
    url_att_cca_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/cca2G', rf, rfa, "url_att_cca5g_smoke")
    url_att_cca_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/cca5G', rf, rfa, "url_att_cca2g_smoke")
    verify_google_ping_from_rg_5g(nvg_599_dut, rf, rfa, "verify_google_ping_from_rg_5g")
    trigger_dfs_channel_change(nvg_599_dut, rf, rfa, 'trigger_dfs_channel_change_from_rg')
    trigger_dfs_channel_change(nvg_599_dut, rf, rfa, 'trigger_dfs_channel_change_from_airties', "ATT_4920_8664D4")
    enable_auto_setup_ssid_via_tr69_cli(nvg_599_dut, '3', rf, rfa, 'conf_auto_setup_ssid_via_tr69_cli', 3)
    enable_auto_setup_ssid_via_tr69_cli(nvg_599_dut, '4', rf, rfa, 'conf_auto_setup_ssid_via_tr69_cli', 3)
    load_airties_firmware(nvg_599_dut, rf, rfa, "load_airties_firmware", "any", "/home/palmer/Downloads/AirTies_Air4920US-AL_FW_3.67.8.3.7623.bin")
    load_airties_firmware(nvg_599_dut, rf, rfa, "load_airties_firmware", "any", "/home/palmer/Downloads/AirTies_Air4920US-AL_FW_3.67.8.3.7623.bin")
    load_airties_firmware(nvg_599_dut, rf, rfa, "load_airties_firmware", "any", "AirTies_Air4920US-AL_FW_2.33.1.2.2112_telnet_enabled_preinstall.bin")
    execute_factory_reset(nvg_599_dut, rf, rfa, "execute_factory_reset")


if send_email == 1:
    nvg_599_dut.email_test_results(rf, nvg_599_dut.software_version)
    exit()

# check this
# load_4920_firmware(nvg_599_dut, rf, rfa, 'load_4920_firmware', 'any', '/home/palmer/Downloads/AirTies_7381.bin')
#   load_airties_firmware(nvg_599_dut, rf, rfa, "load_airties_firmware", "any", "/home/palmer/Downloads/AirTies_Air4920US-AL_FW_3.67.8.3.7623.bin")
#   load_airties_firmware(nvg_599_dut, rf, rfa, "load_airties_firmware", "any", "AirTies_Air4920US-AL_FW_2.33.1.2.2112_telnet_enabled_preinstall.bin")



    # this failed
    # verify_ssid_change_propagated_to_airties(nvg_599_dut,rf, rfa, "verify_ssid_change_propagated_to_airties", "ATT4ujR48sdog","default")

    # this  works
    # nvg_599_dut.set_all_4920s_to_factory_default()
    # the airties we have to have had a prior connection using nmcli
    #  nmcli device wifi connect AirTies_SmartMesh_4PNF password kykfmk8997
    # which would then appear in /etc/NetworManager/system-connections
    # it appears  that all you have to do is "nmcli con up AirTies_SmartMesh_4PNF"
    # this takes the prior active connection down and connects to the new one, which is easier, and then triger
    # on the rg ui from the the wps button
    #
    # def static_reset_4920(ip_4920):
    #     print('In static_reset_4920')
    #     cli_session = pexpect.spawn("telnet " + ip_4920, encoding='utf-8')
    #     cli_session.expect("ogin:")
    # def set_all_4920s_to_factory_default(self):
    #     # show_ip_lan_dict = self.get_rg_sh_ip_lan_info_cli()
    #     show_ip_lan_dict = Nvg599Class.get_rg_ip_lan_info_dict(self)
    # this works----------------------------------------------------------------------------
    # nvg_599_dut.tftp_get_file_cli('LP-PPALMER', 'AirTies_7381.bin', rf, rfa)
    # nvg_599_dut.install_airties_firmware('192.168.1.68', '/home/palmer/Downloads/AirTies_Air4920US-AL_FW_2.49.2.18.7197_FullImage.bin', rf, rfa)
    # nvg_599_dut.install_airties_firmware('192.168.1.68', '/home/palmer/Downloads/AirTies_7381.bin', rf, rfa)
    # nvg_599_dut.install_airties_firmware('192.168.1.68', '/home/palmer/Downloads/airties_telnet_preinstall.bin', rf, rfa)
    # exit()
    # install_airties_firmware(rf, rfa, "install_airties_firmware", '/home/palmer/Downloads/airties_telnet_preinstall.bin', "ATT_4920_8664D4")
    # end this works----------------------------------------------------------------------------

    ## guest_client_cannot_ping_rg(nvg_599_dut, rf, rfa, "guest_client_cannot_ping_rg", "default", "default123")
    ## url_att_friendly_info_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/friendly-info', rf, rfa, 'url_att_friendly_info_smoke')

    ## trigger_dfs_channel_change(nvg_599_dut, rf, rfa, 'trigger_dfs_channel_change', "None")
    # test_speedtest_from_android(nvg_599_dut, 'Galaxy-Note8', test_house_devices_static_info, 'test_speedtest_from_android ', rf, rfa)
    ## ping_gw_from_4920(nvg_599_dut, rf, rfa, "ping_gw_from_4920")
    ## ping_airties_from_rg(nvg_599_dut, rf, rfa, "ping_airties_from_RG")
    ## verify_airties_hello_packet_count_increasing(nvg_599_dut, rf, rfa, "verify_airties_hello_packet_count_increasing")
    ## verify_airties_build_versions(nvg_599_dut, rf, rfa, 'verify_airties_build_versions')
    ## url_att_topology_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/topology', rf, rfa, 'url_att_topology_smoke')
    ## url_att_route_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/route', rf, rfa, 'url_att_route_smoke')
    band5_peers_set_after_airties_association(nvg_599_dut, rf, rfa, "band5_peers_set_after_airties_associationn")
    # nvg_599_dut.tftp_get_file_cli(source_device_name, "nvg599-9.2.2h13d24_1.1.bin","nvg599-9.2.2h13d22_1.1.bin","nvg599-9.2.2h13d20_1.1.bin","nvg599-9.2.2h13d18_1.1.bin","nvg599-9.2.2h13d16_1.1.bin","nvg599-9.2.2h13d14_1.1.bin","nvg599-9.2.2h13d12_1.1.bin","nvg599-9.2.2h13d10_1.1.bin")
    ## url_att_steer_smoke(nvg_599_dut, 'http://192.168.1.254/ATT/steer', rf, rfa, 'url_att_steer_smoke')
    ## test_speedtest_from_android(nvg_599_dut, 'Galaxy-Note8', test_house_devices_static_info, 'test_speedtest_from_android ', rf, rfa)
    ## verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '3', 'ZipKey-PSK', 'Cirrent1',  rf, rfa, "verify_auto_ssid_defaults_via_tr69" )
    ## verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '4', 'ATTPOC', 'Ba1tshop', rf, rfa, "verify_auto_ssid_defaults_via_tr69" )
    ## verify_auto_info_not_present_in_ui(nvg_599_dut, rf, rfa, "verify_auto_info_not_present_in_ui" )
    ## verify_google_ping_from_rg_5g(nvg_599_dut, rf, rfa, "verify_google_ping_from_rg_5g")
    ## local_to_remote_ping(nvg_599_dut,rf, rfa,  '192.168.1.69',  "local_to_remote_ping")

    # before factory reset the order would be
    # reset any airties device to factory defaults and then after reset access the airties default netowrk using
    # the nmcli routines.
    ##  execute_factory_reset(nvg_599_dut, rf, rfa, "execute_factory_reset")
    # nvg_599_dut.enable_guest_network_and_set_passwords(rf, rfa)
    # guest_client_cannot_ping_rg(nvg_599_dut, rf, rfa, "def guest_client_cannot_ping_rg(nvg_599_dut")
    # nvg_599_dut.set_all_4920s_to_factory_default()

if send_email == 1:
    nvg_599_dut.email_test_results(rf, nvg_599_dut.software_version)
    # we want to get the SSID becaause we have to reconnect to it
    # wl_status_info_dict = nvg_599_dut.get_rg_band2_status_cli()
    # rg_ssid = wl_status_info_dict['ssid']
    # print('rg ssid is:'+ str(rg_ssid))

    # then we want to get the specific IP of the named airties device
    # so we would get the dict which is keyed on macs then when we find the name = to the named device we
    # can get the  state and the current ip. If the state is off then the device may be powered off or it is
    #has been factory reset.
    # at that point we have to check the available wireless networks to see if the Airties is in AP mode by
    # looking for its default network.
    # if we find it then we connect to the network as a "sta" enter the basbaglan command,
    # change back to the Rg network and call the click wps method.

exit()

from openpyxl.styles import Color,PatternFill, Font, Border
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
# ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
ws.title = "nvg599 11.5 d1"
ws['A7'] = "Commments"
ws['B7'] = "New"
ws['C7'] = "Test ID "
ws['D7'] = "Test Title"
ws['E7'] = "Purpose/Objective"
ws['F7'] = "Steps"
ws['G7'] = "Results"

ws['A8'] = "Book 1 in Docx File"
ws['A9'] = ""
ws['C10'] = ""



for cell in ws["7:7"]:
    # cell.fill = PatternFill(bgColor = "FEF5A8", fill_type = "solid")
    cell.fill = PatternFill(start_color = "FEF5A8", end_color = "FEF5A8", fill_type = "solid")

    #cell.font = red_font
wb.save('myworkbook.xlsx')
exit()
# wb.save(filename=file)
# >>> source = wb.active
# >>> target = wb.copy_worksheet(source)
# create copies of  a worksheet


#tree = ET.parse('/home/palmer/tmp/config00.xml')
#import xml.etree.ElementTree as ETree
from  xml.etree.ElementTree import  fromstring, ElementTree

#send_email = 1
rf = open('results_file.txt', mode = 'w', encoding = 'utf-8')
rfa  = open('results_file.txt', mode = 'a', encoding = 'utf-8')
#now = datetime.today().strftime("%B %d, %Y,%H:%M")
# rf.write('RG Test run:' + now + '\n')
# rfa.write(now + '\n')
send_email = 1
nvg_599_dut = Nvg599Class()

nvg_599_dut.tftp_rg_firmware_and_install("LP-PPALMER", "nvg599-11.5.0h0d2_1.1.bin", rf, rfa)
exit()
execute_factory_reset(nvg_599_dut, rf, rfa, 'execute_factory_reset')
if send_email == 1:
    nvg_599_dut.email_test_results(rf)
rf.close()
rfa.close()
exit()



#def tftp_get_file_cli(self,remote_file, *source_device_list):
# AirTies_7084.bin
#remote_file = "nvg589-2.2h13d24.bin"
# note that when we re-install the telnet bin, the ip is assumed to be 192.168.2.254  which is what it would be after a factory reset.
# also we assume that we are connected to the AP's default SSID


# remote_file = rg_firmware nvg599-11.5.0h0d1_1.bin"
remote_file = "nvg599-2.2h13d26.bin"
tftp_rg_firmware_and_install(nvg_599_dut, "LP-PPALMER",remote_file, rf,rfa ,"tftp_rg_firmware_and_install")

remote_file = "nvg599-2.2h13d25.bin"
tftp_rg_firmware_and_install(nvg_599_dut, "LP-PPALMER",remote_file, rf,rfa ,"tftp_rg_firmware_and_install")

remote_file = "nvg599-2.2h13d26.bin"
tftp_rg_firmware_and_install(nvg_599_dut, "LP-PPALMER",remote_file, rf,rfa ,"tftp_rg_firmware_and_install")

#tftp_rg_firmware_and_install(nvg_599_dut, "LP-PPALMER","nvg599-11.5.0h0d1_1.1.bin", rf,rfa ,"tftp_rg_firmware_and_install")
#                                                      nvg599-11.5.0h0d1_1.1.bin
#airties_wps_connection_after_airties_factory_reset(nvg_599_dut,"dog", rf, rfa, "airties_wps_connection_after_airties_factory_reset")

exit()


# this looks to be installing into an airties in AP mode
# do I need this, when installing neww firmware or only on factory reset?
nvg_599_dut.install_4920_firmware('192.168.2.254', '/home/palmer/Downloads/airties_telnet_preinstall.bin', rf, rfa)



rf.close()
rfa.close()
exit()
source_device_name = "LP-PPALMER"
remote_file = "airties_telnet_preinstall.bin"
nvg_599_dut.tftp_get_file_cli(source_device_name, "airties_telnet_preinstall.bin")
band5_peers_cleared_after_reset(nvg_599_dut, rf, rfa,"band5_peers_cleared_after_reset")
band5_peers_set_after_airties_association(nvg_599_dut, rf, rfa, "band5_peers_set_after_airties_associationn")
# nvg_599_dut.tftp_get_file_cli(source_device_name, "nvg599-9.2.2h13d24_1.1.bin","nvg599-9.2.2h13d22_1.1.bin","nvg599-9.2.2h13d20_1.1.bin","nvg599-9.2.2h13d18_1.1.bin","nvg599-9.2.2h13d16_1.1.bin","nvg599-9.2.2h13d14_1.1.bin","nvg599-9.2.2h13d12_1.1.bin","nvg599-9.2.2h13d10_1.1.bin")
url_att_steer_smoke(nvg_599_dut,'http://192.168.1.254/ATT/steer',rf,rfa, 'url_att_steer_smoke')
rf.close()
rfa.close()
if send_email == 1:
    nvg_599_dut.email_test_results(rf)
exit()

#execute_factory_reset(nvg_599_dut, rf, rfa, 'execute_factory_reset')
#verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '3', 'ZipKey-PSK', 'Cirrent1',  rf, rfa, "verify_auto_ssid_defaults_via_tr69" )
#verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '4', 'ATTPOC', 'Ba1tshop', rf, rfa, "verify_auto_ssid_defaults_via_tr69" )
#verify_auto_info_not_present_in_ui(nvg_599_dut, rf, rfa, "verify_auto_info_not_present_in_ui" )
verify_google_ping_from_rg_5g(nvg_599_dut, rf, rfa, "verify_google_ping_from_rg_5g")

url_att_topology_smoke(nvg_599_dut,'http://192.168.1.254/ATT/topology',rf, rfa, 'url_att_topology_smoke')
url_att_route_smoke(nvg_599_dut,'http://192.168.1.254/ATT/route',rf,rfa, 'url_att_route_smoke')

#   this should pass
# url_att_friendly_info_smoke(nvg_599_dut,'http://192.168.1.254/ATT/friendly-info',rf,rfa, 'url_att_friendly_info_smoke')
#??????????????????????????????????????????????????


#url_att_cca5g_smoke(nvg_599_dut,'http://192.168.1.254/ATT/cca5g',rf,rfa, 'url_att_cca5g_smoke')
# url_att_steer_smoke(nvg_599_dut,'http://192.168.1.254/ATT/steer',rf,rfa, 'url_att_steer_smoke')
rf.close()
rfa.close()
if send_email == 1:
    nvg_599_dut.email_test_results(rf)
exit()

run_file = open('results_file.txt', mode = 'w', encoding = 'utf-8')
run_file_db  = open('test.txt', mode = 'w', encoding = 'utf-8')
now = datetime.today().strftime("%B %d, %Y,%H:%M")
run_file.write('RG Test run:' + now + '\n')
nvg_599_dut = Nvg599Class()

# peers_xml = band5_att_peers_smoke(nvg_599_dut,'present', run_file, run_file_db, 'band5_att_peers_smoke')

# print('outside of func' + str(peers_xml))
exit()

# tree = ElementTree(fromstring(peers_xml))
# #tree = ETree.fromstring(peers_xml)
# print(str(tree))
# print('1--------------------\n')
# #print(ETree.tostring(tree.getroot()))
# root = tree.getroot()
# print('2---:' +  root.tag + '\n')
# peer_list = []
# for peer in root.iter('peer'):
#     print(str(peer.attrib['address']))
#     peer_list.append(peer.attrib['address'])
# print(peer_list)
# run_file.close()
# run_file_db.close()
# exit()

# parser = ETree.XMLParser(encoding="utf-8")
# tree = ETree.parse("/home/palmer/tmp/config00.xml")
#   root = ET.fromstring(country_data_as_string)
# this routine could use a max lients option
# setup_auto_ssid_via_tr69(nvg_599_dut, '3', '4', rf, rfa, "setup_auto_ssid_via_tr69")
# setup_auto_ssid_via_tr69(nvg_599_dut, '4', rf, rfa, "setup_auto_ssid_via_tr69")
# print(ETree.tostring(tree.getroot()))
# mytreeroot = tree.getroot()
# mytree = ETree.tostring(tree)
# print('1--------------------\n')
# print(ETree.tostring(tree.getroot()))
# root = tree.getroot()
# print('2---:' +  root.tag + '\n')
# peer_list = []
# for peer in root.iter('peer'):
#     print(str(peer.attrib['address']))
#     peer_list.append(peer.attrib['address'])
# print(peer_list)
# exit()



band_5_wl_status_dict = nvg_599_dut.get_rg_band5_status_cli()
print('---------------band_5_wl_status_dict:' + str(band_5_wl_status_dict['bssid']))
print('---------------band_5_wl_st channel:' + str(band_5_wl_status_dict['channel']))
print('---------------band_5_wl_stat--bw:' + str(band_5_wl_status_dict['bandwidth']))


exit()
# set the file at to the file to use
# source_device_name = "LP-PPALMER"
# nvg_599_dut.tftp_get_file_cli(source_device_name, "nvg599-9.2.2h13d26_1.1.bin","nvg589-9.2.2h13d26_1.1.bin")
#
# test_rg_upgrade(nvg_599_dut, '/home/palmer/Downloads/nvg599-9.2.2h13d26_1.1.bin', rf, rfa)
execute_factory_reset(nvg_599_dut, rf, rfa, 'execute_factory_reset')
verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '3', 'ZipKey-PSK', 'Cirrent1',  rf, rfa, "verify_auto_ssid_defaults_via_tr69" )
verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '4', 'ATTPOC', 'Ba1tshop', rf, rfa, "verify_auto_ssid_defaults_via_tr69" )
verify_auto_info_not_present_in_ui(nvg_599_dut, rf, rfa, "verify_auto_info_not_present_in_ui" )
verify_google_ping_from_rg_5g(nvg_599_dut, rf, rfa, "verify_google_ping_from_rg_5g")

# need to make sure 4920s are associated first
ping_gw_from_4920(nvg_599_dut,rf,rfa, "ping_gw_from_4920")
ping_airties_from_rg(nvg_599_dut, rf, rfa, "ping_airties_from_RG")
verify_airties_hello_packet_count_increasing(nvg_599_dut, rf, rfa, "verify_airties_hello_packet_count_increasing")
verify_airties_build_versions(nvg_599_dut, rf, rfa, 'verify_airties_build_versions')

#airties_ap_net = "AirTies_SmartMesh_4PNF"
#nvg_599_dut.wps_pair_default_airties(airties_ap_net)

rf.close()
rfa.close()
if send_email == 1:
    nvg_599_dut.email_test_results(rf)
exit()


#def tftp_get_file_cli(self,remote_file, *source_device_list):

#remote_file = "nvg589-2.2h13d24.bin"
#source_device_name = "LP-PPALMER"
#remote_file = "a1.txt"
#nvg_599_dut.tftp_get_file_cli(source_device_name, "nvg599-9.2.2h13d24_1.1.bin","nvg599-9.2.2h13d22_1.1.bin","nvg599-9.2.2h13d20_1.1.bin","nvg599-9.2.2h13d18_1.1.bin","nvg599-9.2.2h13d16_1.1.bin","nvg599-9.2.2h13d14_1.1.bin","nvg599-9.2.2h13d12_1.1.bin","nvg599-9.2.2h13d10_1.1.bin")

# we first get the IP of an associated 4920.
# if none availabe then exit with fail message
# # self.ip_lan_connections_dict_cli[connected_device_mac] = {}
# ip_lan_connections_dict_cli[connected_device_mac] = {}
# ip_lan_connections_dict_cli[connected_device_mac]["IP"] = connected_device_ip
# ip_lan_connections_dict_cli[connected_device_mac]["Name"] = connected_device_name
# ip_lan_connections_dict_cli[connected_device_mac]["State"] = connected_device_status
# ip_lan_connections_dict_cli[connected_device_mac]["DHCP"] = connected_device_dhcp
# ip_lan_connections_dict_cli[connected_device_mac]["Port"] = connected_device_port
# verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '3', rf, rfa, "verify_auto_ssid_defaults_via_tr69" )
# print(' verify_auto_ssid_defaults_via_tr69 ssid:'  + status)
######################################################### end test area, clean up the junk below
# set the file at to the file to use
# test_rg_upgrade(nvg_599_dut, '/home/palmer/Downloads/nvg599-9.2.2h13d25_1.1.bin', rf, rfa)

verify_auto_ssid_defaults_via_tr69(nvg_599_dut, '4', rf, rfa, "verify_auto_ssid_defaults_via_tr69")
#print('verify_auto_ssid_defaults_via_tr69 ssid:' + status)

setup_auto_ssid_via_tr69(nvg_599_dut, '3', rf, rfa, "setup_auto_ssid_via_tr69")

setup_auto_ssid_via_tr69(nvg_599_dut, '4', rf, rfa, "setup_auto_ssid_via_tr69")

verify_auto_info_not_present_in_ui(nvg_599_dut, rf, rfa, "verify_auto_info_not_present_in_ui")
#status = nvg_599_dut.set_auto_setup_ssid_via_tr69_cli(nvg_599_dut, "3", rf, rfa, 'nvg_599_dut.set_auto_setup_ssid_via_tr69_cli')
# status = nvg_599_dut.set_auto_setup_ssid_via_tr69_cli("4", rf, rfa)

#status = nvg_599_dut.set_auto_setup_ssid_via_tr69_cli(nvg_599_dut, "4", rf, rfa, 'nvg_599_dut.set_auto_setup_ssid_via_tr69_cli')



######################################################################################################
######################  Experimental code  below #####################
nmcli_connection = "Wired"
nvg_599_dut.nmcli_set_connection(nmcli_connection, "down")
sleep(5)
nvg_599_dut.nmcli_set_connection(nmcli_connection, "up")


now = datetime.today().strftime("%B %d, %Y,%H:%M")
rf = open('results_file.txt', mode = 'w', encoding = 'utf-8')
rfa  = open('results_file.txt', mode = 'a', encoding = 'utf-8')
rf.write(now + '\n')
rfa.write(now + '\n')
rf.write('Test title:test_auto_ssid_3_default_tr69_values \n')
test_status = test_auto_ssid_default_tr69_values(nvg_599_dut,"3",rf,rfa)
rf.write('    ' + test_status + '\n')
print('    Test Status:' + test_status + '\n')
rf.write('\n')
print('\n')

rf.write('Test title:test_auto_ssid_4_default_tr69_values \n')
test_status = test_auto_ssid_default_tr69_values(nvg_599_dut,"4",rf,rfa)
rf.write('    ' + test_status + '\n')
print('    Test Status:' + test_status + '\n')
rf.write('\n')
print('\n')

rf.write('Test title:test_verify_auto_ssid_info_not_present_in_ui \n')
test_status = verify_auto_info_not_present_in_ui(nvg_599_dut,rf,rfa)
rf.write(test_status + '\n')

rf.close()
rfa.close()
sleep(20)
nvg_599_dut.email_test_results(rf)
exit()

status_page = nvg_599_dut.get_ui_home_network_status_page()
print(status_page)
exit()
#
# default_tr69_auto_ssid_values = nvg_599_dut.get_tr69_auto_ssid(ssid)
# # print(default_tr69_auto_ssid_values)
# #exit()
# if (default_tr69_auto_ssid_values.find('.' + ssid + '.Enable 0') != -1):
#     print('Pass ssid:' + ssid + ' default set to 0')
# else:
#     print('Fail ssid:' + ssid + ' default not set to 0')
#
# exit()

now = datetime.today().strftime("%B %d, %Y,%H:%M")
results_file = open('results_file.txt', mode = 'w', encoding = 'utf-8')

rf = open('results_file.txt', mode = 'w', encoding = 'utf-8')
rfa  = open('results_file.txt', mode = 'a', encoding = 'utf-8')
rf.write(now + '\n')
rfa.write(now + '\n')

results_file_archive=open('results_file.txt', mode ='a', encoding='utf-8')
results_file.write(now + '\n')
results_file_archive.write(now + '\n')

nvg_599_dut = Nvg599Class()

upgrade_rg_file ='/home/palmer/Downloads/nvg599-9.2.2h13d26_1.1.bin'
#test_status, duration = nvg_599_dut.upgrade_rg(upgrade_rg_file, rf, rfa)
test_status, duration = nvg_599_dut.upgrade_rg('/home/palmer/Downloads/nvg599-9.2.2h13d26_1.1.bin', rf, rfa)

sleep(300)
results_file.write("Test Title: RG Upgrade :" + upgrade_rg_file + " Test case " + test_status  + "Duration:" + duration  +'\n')
results_file_archive.write("Test Title: RG Upgrade :" + upgrade_rg_file + " Test case " + test_status  + "Duration:" + duration  +'\n')
#results_file_archive.write("Test Title: RG Upgrade:" + upgrade_rg_file + " Pass " + '\n')
# nvg_599_dut.factory_reset_rg()
#
# ssid = 3
# nvg_599_dut.set_auto_setup_ssid_via_tr69_cli(ssid)
# results_file.write("Test Title: Auto SSID 3 setup: Pass")
# results_file_archive.write("Test Title: Auto SSID 3 setup: Pass")
#
# ssid = 4
# nvg_599_dut.set_auto_setup_ssid_via_tr69_cli(ssid)
# results_file.write("Test Title: Auto SSID 4 setup: Pass")
# results_file_archive.write("Test Title: Auto SSID 4 setup: Pass")
# results_file.write("Test Title: RG Factory Reset:" + " Pass " + '\n')
# results_file_archive.write("Test Title: RG Factory Reset:" + " Pass " + '\n')
#
# results_file.write("Test Title: RG Factory Reset: turn_off_supplicant_cli(): Pass \n")
# results_file_archive.write("Test Title: RG Factory Reset: turn_off_supplicant_cli(): Pass \n")
#
# results_file.write("Test Title: RG Factory Reset: enable_sshd_ssh_cli(): Pass \n")
# results_file_archive.write("Test Title: RG Factory Reset: enable_sshd_ssh_cli(): Pass \n")
#
# results_file.write("Test Title: RG Factory Reset: conf_tr69_eco_url(): Pass \n")
# results_file_archive.write("Test Title: RG Factory Reset: conf_tr69_eco_url(): Pass \n")
#
# results_file.write("Test Title: RG Factory Reset: turn_off_wi_fi_security_protection_cli(): Pass \n")
# results_file_archive.write("Test Title: RG Factory Reset: turn_off_wi_fi_security_protection_cli(): Pass \n")
#
# results_file.write("Test Title: RG Factory Reset: enable_parental_control():Pass \n")
# results_file_archive.write("Test Title: RG Factory Reset: enable_parental_control(): \n")

# start = time.time()
# sleep(10)
# end = time.time()
#
# # round_dog = round(start)
# #print('time: ' + str(time.time()))
# print('time: ' + str(round(end - start)))

#upgrade_rg_file ='/home/palmer/Downloads/nvg599-9.2.2h13d22_1.1.bin'
# upgrade_rg_file ='/home/palmer/Downloads/nvg599-9.2.2h12d15_1.1.bin'
# upgrade_rg_file ='/home/palmer/Downloads/nvg599-9.2.2h2d23_1.1.bin'

# airties_ap_net = "AirTies_Air4920_33N3"
# nvg_599_dut.wps_pair_default_airties(airties_ap_net)
exit()
# nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, 36)
# this will go in factory reset once ribust enough
airties_ap_net = "AirTies_SmartMesh_4PNF"
nvg_599_dut.wps_pair_default_airties(airties_ap_net)
sleep(300)
airties_ap_net = "AirTies_Air4920_33N3"
nvg_599_dut.wps_pair_default_airties(airties_ap_net)
# wps_pair_default_airties(airties_ap_net)

# nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, 100)
nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, 36)

exit()
upgrade_rg_file ='/home/palmer/Downloads/nvg599-9.2.2h13d17_1.1.bin'
nvg_599_dut.upgrade_rg(upgrade_rg_file)
sleep(300)
nvg_599_dut.factory_reset_rg()
sleep(300)
nvg_599_dut.enable_guest_network_and_set_password_ssid()

#nvg_599_dut.enable_guest_network_and_set_password_ssid()
rg_url = 'http://192.168.1.254/'
nvg_599_dut.wps_button_click(rg_url)
airties_ap_net = "AirTies_SmartMesh_4PNF"
nvg_599_dut.wps_pair_default_airties(airties_ap_net)

sleep(300)
nvg_599_dut.wps_button_click(rg_url)
airties_ap_net = "AirTies_Air4920_33N3"

# nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, 100)
nvg_599_dut.ui_set_band_bandwith_channel('5g', 80, 36)


# test_dfs(nvg_599_dut)
# tst_ping_rg_power_level(nvg_599_dut, '192.168.1.94', '20')
# nvg_599_dut.ui_enable_guest_network_and_set_password_ssid()
# exit()
# wifi_password_script = "1111111111"
# custom_security = "Custom Password"
# set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, wifi_password_script)
# print('******** set_wifi_return_code: ' + set_wifi_return_code)
# min_ping, avg_ping, max_ping, mdev_ping, sent, received, loss  = nvg_599_dut.ping_from_local_host(remote_ip, number_of_pings)
# print('min:'+ min_ping + ' max:' + max_ping)
ping_file = open('ping_file_with_power_change_test.txt', 'a')
# now = datetime.today().isoformat()
now = datetime.today().strftime("%B %d, %Y,%H:%M")
# ping_file.writelines('Ping ' + remote_ip + ' RG Pwr %:' + str(percentage) +  ' Sent:' + sent + ' Received:' + received + ' Percent loss:' + loss + '%' + ' Date:' + now + '\n')
ping_file.writelines('Date:' + now + '599 FW Ver:' + nvg_599_dut.software_version + ' Ser. No:' + nvg_599_dut.serial_number + '\n')
# ping_file.writelines('Ping ' + remote_ip + ' RG Pwr %:' + str(percentage) +  ' Sent:' + sent + ' Received:' + received + ' Percent loss:' + loss + '%\n')
# ping_file.writelines('Date:' + now + ' Ping ' + remote_ip + ' RG Pwr %:' + str(percentage) +  ' Sent:' + sent + ' Received:' + received + ' Percent loss:' + loss + '%\n')

# ping_file.writelines("599 FW Ver:" + nvg_599_dut.software_version + " Ser. No:" +
#                      nvg_599_dut.serial_number + '  min_ping:' + min_ping + '  avg_ping:' + avg_ping +
#                      '  max_ping:' + max_ping + '  max dev:' + mdev_ping +'\n')

# ping_file.writelines('Minimum:' + min_ping + ' Average::' + avg_ping + ' Maximum:' + max_ping + ' max dev:' + mdev_ping +'\n')

# ping_file.writelines('Date:' + now + " 599 FW Ver:" + nvg_599_dut.software_version + " Ser. No:" +
#                      nvg_599_dut.serial_number + '  min_ping:' + min_ping + '  avg_ping:' + avg_ping +
#                      '  max_ping:' + max_ping + '  max dev:' + mdev_ping)
ping_file.writelines('\n')
ping_file.close()
Nvg599Class.nmcli_test()
exit()
########################################################################################################3
# evice_dict = nvg_599_dut.ui_get_device_list()
# print('device_dict:', device_dict)
# nvg_599_dut.disable_enable_wifi_2_4_and_5g_wifi()
# exit()
# nvg_599_dut.enable_parental_control()
# nvg_599_dut.enable_guest_network_and_set_password_ssid()
# nvg_599_dut.factory_reset_rg()
# nvg_599_dut.turn_off_supplicant_cli()
# nvg_599_dut.set_fixed_ip_allocation()
# nvg_599_dut.enable_guest_network_and_set_password_ssid()
# Nvg599Class().run_speed_test_from_android_termux("192.168.1.70")
# test_rg_upgrade_speedtest(nvg_599_dut,firmware_599_available,firmware_599_names)


# import datetime
# date = datetime.datetime.strptime(date, "%m%d%Y")
# from pylab import *

speed_test_result_list = []
# download, upload = Nvg599Class.run_speed_test_from_android_termux("192.168.1.70")
for i in  range(4):
    print('loop:' + str(i))
    speed_test_result_tuple  = Nvg599Class.run_speed_test_from_android_termux("192.168.1.70")
    speed_test_result_list.append(speed_test_result_tuple)

down_load = []
up_load = []

for a,b in speed_test_result_list:
    print('download speed: ' + a + ' Upload speed:' + b)
    down_load.append(a)
    up_load.append(b)

import matplotlib.pyplot as plt
# nvg_599_dut.enable_sshd_ssh_cli()
# nvg_599_dut.turn_off_supplicant_cli()
# nvg_599_dut.conf_tr69_eco_url()
# url_to_check = "http://192.168.1.254/cgi-bin/home.ha"
# nvg_599_dut.factory_reset_rg()
# nvg_599_dut.enable_parental_control()
# nvg_599_dut.turn_off_wi_fi_security_protection_cli()
# dfs_file = open('dfs_file.txt','a')
# nvg_599_dut.factory_reset_rg()

# the default  url is "http://192.168.1.254/cgi-bin/home.ha"
# nvg_599_dut.factory_reset_rg()
# security, current_password = nvg_599_dut.ui_get_wifi_password()
# print('current pasword:' + current_password)
# print('security:' + security + ' new pasword:' + current_password)
# wpa or defwpa
default_security = "Default Password"
custom_security = "Custom Password"

# too long >  63
wifi_password_script = "1111111111"
set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, wifi_password_script)
print('******** set_wifi_return_code: ' + set_wifi_return_code)

password_too_short_7 = "1234567"
password_max_63 = "000000000011111111112222222222333333333344444444445555555555666"
password_too_long_64 = "0000000000111111111122222222223333333333444444444455555555556666"
password_too_long_65 = "00000000001111111111222222222233333333334444444444555555555566666"
password_too_long_66 = "000000000011111111112222222222333333333344444444445555555555666666"
password_too_long_67 = "0000000000111111111122222222223333333333444444444455555555556666666"
password_too_long_68 = "00000000001111111111222222222233333333334444444444555555555566666666"
password_too_long_69 = "000000000011111111112222222222333333333344444444445555555555666666666"
password_too_long_70 = "0000000000111111111122222222223333333333444444444455555555556666666666"
password_too_long_71 = "00000000001111111111222222222233333333334444444444555555555566666666666"

# set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, password_too_long_64)
# print('******** set_wifi_return_code: ' + set_wifi_return_code)
# print('############################################')
# set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, password_too_long_65)
# print('******** set_wifi_return_code: ' + set_wifi_return_code)
# print('############################################')
# set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, password_too_long_66)
# print('******** set_wifi_return_code: ' + set_wifi_return_code)
# print('############################################')
# set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, password_too_long_67)
# print('******** set_wifi_return_code: ' + set_wifi_return_code)
# print('############################################')
# set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, password_too_long_68)
# print('******** set_wifi_return_code: ' + set_wifi_return_code)
# print('############################################')
# set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, password_too_long_69)
# print('******** set_wifi_return_code: ' + set_wifi_return_code)
# print('############################################')
# set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, password_too_long_70)
# print('******** set_wifi_return_code: ' + set_wifi_return_code)
# print('############################################')
# set_wifi_return_code = nvg_599_dut.ui_set_wifi_password(custom_security, password_too_long_71)
# print('******** set_wifi_return_code: ' + set_wifi_return_code)
# print('############################################')
# exit()

# min chars is 8
#
# # max chars is 63
# #password_just_right ="000000000011111111112222222222333333333344444444445555555555666"
# update_rg ='/home/palmer/Downloads/nvg599-9.2.2h13d10_1.1.bin'
# nvg_599_dut.update_rg(update_rg)
# exit()
#

ALL_BAND5_CHANNELS = [104, 108, 112, 116, 132, 136, 140, 144, 149, 153, 157, 161, 165]
#ALL_BAND5_CHANNELS = [36, 40, 44, 48,52, 56, 60, 64, 100, 104, 108, 112, 116, 132, 136, 140, 144, 149, 153, 157, 161, 165]
ALL_BAND5_BANDWIDTHS= [20,40,80]
xALL_BAND5_CHANNELS = [52, 56]
xALL_BAND5_BANDWIDTHS= [20, 40]


# ui_set_band_bandwith_channel(self, band, bandwidth, channel):
channel_band_file = open('channel_band_file_to_4920.txt', 'a')
now = datetime.today().isoformat()
channel_band_file.write("\" + ""Test Title:Channel/Channel band test:" +"\n")
channel_band_file.write(now + "\n")
software_version = nvg_599_dut.software_version
channel_band_file.write("NVG599 Firmware:"+ software_version + "\n")
for band5_channel in range(len(ALL_BAND5_CHANNELS)):
    for band5_bandwidth in range(len(ALL_BAND5_BANDWIDTHS)):
        print(' ')
        print('band bandwidth channel: ' + str(ALL_BAND5_CHANNELS[band5_channel]) + "  "  + str(ALL_BAND5_BANDWIDTHS[band5_bandwidth]))
        print(' ')
        sleep(30)
        nvg_599_dut.ui_set_band_bandwith_channel('5g', ALL_BAND5_BANDWIDTHS[band5_bandwidth], ALL_BAND5_CHANNELS[band5_channel])

        sleep(60)
        ping_results = nvg_599_dut.ping_check('192.168.1.80')
        print('Channel:' + str(ALL_BAND5_CHANNELS[band5_channel]) + " Bandwidth:"  + str(ALL_BAND5_BANDWIDTHS[band5_bandwidth]) + ' Ping result:' + str(ping_results))
        #min_ping, avg_ping, max_ping, mdev_ping = nvg_599_dut.ping_from_local_host('192.168.1.77')
        #print('min_ping:'+  ping_file = open('ping_file.txt', 'a') min_ping + ' avg_ping:' + avg_ping + ' max_ping:' + max_ping)

        channel_band_file.writelines('Channel:' + str(ALL_BAND5_CHANNELS[band5_channel]) + " Bandwidth:"  + str(ALL_BAND5_BANDWIDTHS[band5_bandwidth]) + ' Ping result:' + str(ping_results) + '\n')
        #                     self.serial_number + '  min_ping:' + min_ping + '  avg_ping:' +
        #                     '  max_ping:' + max_ping + '  max dev:' + mdev_ping)
        #ping_file.writelines('\n')
channel_band_file.close()
# nvg_599_dut.ui_set_band_bandwith_channel('5g', band5_bandwidth, band5_channel)

exit()
# remote_driver = nvg_599_dut.remote_webserver()
# remote_driver.get("http://www.firefox.com")
# nvg_599_dut.enable_sshd_ssh_cli()
# nvg_599_dut.turn_off_supplicant_cli()
# nvg_599_dut.conf_tr69_eco_url()
# url_to_check = "http://192.168.1.254/cgi-bin/home.ha"
# nvg_599_dut.factory_reset_rg(url_to_check)
# nvg_599_dut.enable_parental_control()
# nvg_599_dut.turn_off_wi_fi_security_protection_cli()
# dfs_file = open('dfs_file.txt','a')
# test_dfs(nvg_599_dut,dfs_file)
# Nvg599Class.factory_test()
# dfs_file.close()
# nvg_599_dut.poc_for_youtube()
exit()

ssid_band5 = "ATTqbrAnYs"
ssid_guest = "ATTqbrAnYs_Guest"
ssid_band2 = "ATTqbrAnYs"

nvg_599_dut.set_ui_ssid(ssid_band5,ssid_guest,ssid_band2)

band2, guest, band5 = nvg_599_dut.get_ui_ssid()
print('band 2: '+ band2 + '  guest: '+ guest + '  band5: ' + band5)

exit()
show_ip_lan_dict = nvg_599_dut.get_rg_ip_lan_info_dict()

print(type(show_ip_lan_dict))
pprint.pprint(show_ip_lan_dict, width = 1)
exit()
#down_load_speed, up_load_speed = nvg_599_dut.run_speed_test_cli(test_ip)

#url_to_check = "http://192.168.1.254/cgi-bin/home.ha"

nvg_599_dut.factory_reset_rg(rf, rfa)
# These are now part of standard factory reset.
# nvg_599_dut.enable_sshd_ssh_cli()
# nvg_599_dut.conf_tr69_eco_url()
# nvg_599_dut.turn_off_supplicant_cli()
# nvg_599_dut.ping_from_local_host('192.1681.228')

print('1' + str(nvg_599_dut.get_airties_ip('airties_1_5g')))
print('2' + str(nvg_599_dut.get_airties_ip('airties_2_5g')))
print('3' + str(nvg_599_dut.get_airties_ip('airties_3_5g')))

results_str ="1234 abcd \n 1234 5678"
nvg_599_dut.email_test_results(results_str)

# ui_dict = nvg_599_dut.ui_get_device_list()



#for key in ui_dict:
#    print(key)

exit()

# tst_ping(nvg_599_dut,results_file,"192.168.1.239")

# tst_speed_test(nvg_599_dut,results_file,"192.168.1.239")

# dfs_file = open('dfs_file.txt','a')
# test_dfs(nvg_599_dut,dfs_file)
# dfs_file.close()

# nvg_599_dut.get_rg_sh_ip_lan_info_cli()

# Nvg599Class.get_rg_sh_ip_lan_info_cli('arris-Latitude-MBR')

# results_str = open('results_file.txt','r').read()
# nvg_599_dut.email_test_results(results_str)
# exit()

#print('dog' + str(DFS_CHANNELS))
#print ('dict' + str(test_house_devices_static_info))
#exit()
#nvg_599_dut.cli_sh_wi_all_clients()
#test_ping_device_name('arris-Latitude-MBR')

# print(mo1)
# print ('model ', mo1.group(1))

# print ('Serial Number', mo1.group(2))
# print ('Uptime ', mo1.group(3))

 # exit()

# rgModel= mo1.group(1)
# serialNumber= mo1.group(2)
# addition

# if rgModel=='NVG599':
#    print('we are going to instantiate an NVG599')
#    nvg599DUT = Nvg599Class()
#    nvg599DUT.printme()
#    nvg599DUT.turnOffSupplicant()


# else:
#    print('what  to instantiate an NVG599')
#
# exit()


# airTiesTerm = pexpect.spawn("telnet 192.168.1.67")
# sleep(1)
# airTiesTerm.expect("ogin:")
# airTiesTerm.sendline('root')
# airTiesTerm.expect("#")

# airTiesTerm.sendline('wl -i wl0 chanspec')
# airTiesTerm.expect("#")
# airTies2GResult=airTiesTerm.before
# print(airTies2GResult)
# airTies2GChannel = airTies2GResult.split()[-2]
# print("2G",airTies2GChannel)
# print("")
# airTiesTerm.sendline('wl -i wl1 chanspec')
# airTiesTerm.expect("#")
# airTies5GResult=airTiesTerm.before
# airTies5GChannelInfo = airTies5GResult.split()[-2]
# print("5GInfo",airTies5GChannelInfo)
# airTies5GChannel =airTies5GChannelInfo.split("/")[0]
# print("5G Channel,airTies5GChannel)
# airTies5GBandwidth =airTies5GChannelInfo.split("/")[1]
# print("5G Bandwidth",airTies5GBandwidth)
# resultFile.close()
print("###############################################################################")
print("###############################################################################")


list2G = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
list5G = [36, 40, 44, 48, 52, 56, 60, 64, 100, 104, 108, 112, 116, 1132, 136, 140, 144, 149, 153, 161]
list2GLite = [1]
list5GLite = [36]

airTiesTerm = pexpect.spawn("telnet 192.168.1.67",encoding='utf-8')
airTiesTerm.expect("ogin:")
airTiesTerm.sendline('root')
airTiesTerm.expect("#")

airTiesTerm.logfile= sys.stdout
# channelResultFP = open('channelResult.txt', 'w+')
# for l2g in list2GLite:
#     for i in list5GLite:
#         print("Config RG channel="+str(i))
#         output = rgTerm.sendline('tr69 SetParameterValues InternetGatewayDevice.LANDevice.1.WLANConfiguration.5.Channel=' + str(i))
#         rgTerm.expect("UNLOCKED>")
#         # re-telnet into the airTies each time because the telnet session gets hung when channel is reset
#         print("Checking airTies for channel match: "+str(i))
#         # I think we want to wait before trying to login to the airtiesssh-add ~/.ssh/id_rsa
#         sleep(200)
#         airTiesTerm = pexpect.spawn("telnet 192.168.1.67", encoding='utf-8')
#         sleep(1)
#         airTiesTerm.expect("ogin")
#         airTiesTerm.sendline('root')
#         airTiesTerm.expect("#")
#         airTiesTerm.sendline('wl -i wl0 chanspec')
#         airTiesTerm.expect("#")
#         airTies2GResult = airTiesTerm.before
#         print("airties result " + str(airTies2GResult))
#         airTies2GChannel = airTies2GResult.split()[-2]
#         airTies2GChannel = airTies2GChannel.split()[-1]
#
#         print("AirTies 2G = ", airTies2GChannel)
#         print("")
#         sleep(2)
#
#         # airTiesTerm.sendline('wl -i wl1 status')   // we want this
#         airTiesTerm.sendline('wl -i wl1 chanspec')
#         airTiesTerm.expect("#")
#         airTies5GResult = airTiesTerm.before
#         airTies5GChannelInfo = airTies5GResult.split()[-2]
#         print("5G Info ", airTies5GChannelInfo)
#
#         sleep(2)
#         #airTies5GChannel = airTies5GChannelInfo.split("/")[-2]
#         airTies5GChannel = airTies5GChannelInfo.split("/")[0]
#
#         print("5G Channel = ", airTies5GChannel)
#         airTies5GBandWidth = airTies5GChannelInfo.split("/")[-1]  # type: object
#         print("5G Bandwidth = ", airTies5GBandWidth)
#
#         airTiesTerm.sendline('exit')
#         airTiesTerm.sendline("\x1b\r")
#         airTiesTerm.terminate(force=True)
#
#         sleep(10)
#
#         #print (type(i))
#         #print ("airTies5GChannel is of type "+ type(airTies5GChannel))
#
#
#         if i == int(airTies5GChannel):
#             channelResultFP.write(" \n")
#             channelResultFP.write(" 2G = "+ str(l2g) + " RG channel= "+ str(i) + " airTies5GChannel = " + airTies5GChannel  +  "  Passed" )
#             #channelResultFP.write("\r\n 2G = "+ str(l2g) + " RG channel= "+ str(i) + " airTies5GChannel = " + airTies5GChannel  +  "Passed" + "\r\n")
#
#             channelResultFP.write("\n ")
#
#             print(">> 2G = " + str(l2g) + " RG channel= "+ str(i) + " airTies5GChannel = " + airTies5GChannel  +  " Passed")
#             print ("----l2g--------------------------------------------- -------------------")
#         else:
#             channelResultFP.write(" ")
#             #channelResultFP.write("\r\n 2G = " + str(l2g) + "RG channel= "+ str(i) + " airTies5GChannel = " + airTies5GChannel  +  " Failed" + "\r\n")
#             channelResultFP.write(" 2G = " + str(l2g) + "RG channel= "+ str(i) + " airTies5GChannel = " + airTies5GChannel  +  " Failed  ")
#
#             channelResultFP.write(" ")
#
#         sleep(30)
#
#
#
#
# ### this used to work
# channelResultFP.seek(0)
# channelResultContents = channelResultFP.read()
# channelResultFP.close()

# resultFile.close()
# sleep(1)

# Note that, for Python 3 compatibility reasons, we are using spawnu and
# importing unicode_literals (above). spawnu accepts Unicode input and
# unicode_literals makes all string literals in this script Unicode by default.
p = pexpect.spawnu('uptime')

# This parses uptime output into the major groups using regex group matching.
p.expect(
    r'up\s+(.*?),\s+([0-9]+) users?,\s+load averages?: ([0-9]+\.[0-9][0-9]),?\s+([0-9]+\.[0-9][0-9]),?\s+([0-9]+\.[0-9][0-9])')
duration, users, av1, av5, av15 = p.match.groups()

# The duration is a little harder to parse because of all the different
# styles of uptime. I'm sure there is a way to do this all at once with
# one single regex, but I bet it would be hard to read and maintain.
# If anyone wants to send me a version using a single regex I'd be happy to see it.
days = '0'
hours = '0'
mins = '0'
if 'day' in duration:
    p.match = re.search(r'([0-9]+)\s+day', duration)
    days = str(int(p.match.group(1)))
if ':' in duration:
    p.match = re.search('([0-9]+):([0-9]+)', duration)
    hours = str(int(p.match.group(1)))
    mins = str(int(p.match.group(2)))
if 'min' in duration:
    p.match = re.search(r'([0-9]+)\s+min', duration)
    mins = str(int(p.match.group(1)))

# Print the parsed fields in CSV format.
# print('days, hours, minutes, users, cpu avg 1 min, cpu avg 5 min, cpu avg 15 min')
# print('%s, %s, %s, %s, %s, %s, %s' % (days, hours, mins, users, av1, av5, av15))
exit()

from time import sleep
from selenium import webdriver

# driver = webdriverhttps://www.waketech.edu/programs-courses/credit/electrical-systems-technology/degrees-pathways.Chrome('/usr/local/bin/chromedriver')
driver = webdriver.Firefox(executable_path='/usr/local/bin/geckodriver')
# driver.get('http://www.google.com')
driver.get('http://192.168.1.254/cgi-bin/sysinfo.ha')
driver.implicitly_wait(20)
# driver.find_elements_by_tag_name("Settings") // this is for 599
driver.find_element_by_link_text("Settings").click()

# driver.findElement(By.linkText("Home Network")).click()
driver.implicitly_wait(20)

driver.find_element_by_link_text("LAN").click()
driver.implicitly_wait(20)
# driver.maximize_window()

driver.find_element_by_link_text("Wi-Fi").click()
driver.implicitly_wait(20)

password = driver.find_element_by_id("ADM_PASSWORD")

password.send_keys("8>1769&295")

driver.find_element_by_class_name('button').click()
# driver.find_element_by_xpath("//button[@value='Submit']").click()
# button.click()

driver.implicitly_wait(20)

# ENter device access code

sleep(30)
driver.quit()

child = pexpect.spawn("ssh root@192.168.1.254")
sleep(1)
# exit()

from time import sleep
from selenium import webdriver
import pexpect
import re
#import serial

# driver = webdriverhttps://www.waketech.edu/programs-courses/credit/electrical-systems-technology/degrees-pathways.
# Chrome('/usr/local/bin/chromedriver')
driver = webdriver.Firefox(executable_path='/usr/local/bin/geckodriver')
# driver.get('http://www.google.com')
driver.get('http://192.168.1.254')
driver.implicitly_wait(20)
# driver.find_elements_by_tag_name("Settings") // this is for 599
driver.find_element_by_link_text("Settings").click()

# driver.findElement(By.linkText("Home Network")).click()
driver.implicitly_wait(20)

driver.find_element_by_link_text("LAN").click()
driver.implicitly_wait(20)
# driver.maximize_window()

driver.find_element_by_link_text("Wi-Fi").click()
driver.implicitly_wait(20)

password = driver.find_element_by_id("ADM_PASSWORD")

password.send_keys("8>1769&295")

driver.find_element_by_class_name('button').click()
# driver.find_element_by_xpath("//button[@value='Submit']").click()
# button.click()

driver.implicitly_wait(20)

# ENter device access code

sleep(30)
driver.quit()

child = pexpect.spawn("ssh root@192.168.1.254")
sleep(1)
# exit()

child.expect("assword:")
child.sendline('alcatel')
# child.sendline('*<#/53#1/2')
print("logged in to host")
child.expect("#")
child.sendline('exit')
sleep(5)
# driver.webdriver.quit()
exit()
# driver = webdriver.firefox('/home/palmer/.local/lib/python2.7/site-packages/chromedriver')
# driver = webdriver.chrome()
# browser.get('https://www.google.com')
days = '0'
hours = '0'
mins = '0'
if 'day' in duration:
    p.match = re.search(r'([0-9]+)\s+day', duration)
    days = str(int(p.match.group(1)))
if ':' in duration:
    p.match = re.search('([0-9]+):([0-9]+)', duration)
    hours = str(int(p.match.group(1)))
    mins = str(int(p.match.group(2)))
if 'min' in duration:
    p.match = re.search(r'([0-9]+)\s+min', duration)
    mins = str(int(p.match.group(1)))

# Print the parsed fields in CSV format.
# print('days, hours, minutes, users, cpu avg 1 min, cpu avg 5 min, cpu avg 15 min')
# print('%s, %s, %s, %s, %s, %s, %s' % (days, hours, mins, users, av1, av5, av15))
print("Turning off supplicant")
ip = "192.168.1.254"
password = '<#/53#1/2'
child = pexpect.spawn("telnet " "192.168.1.254")
sleep(1)
# exit()
child.expect("ogin:")
child.sendline("admin")

child.expect("assword:")
child.sendline('*<#/53#1/2')
child.expect(">")

child.sendline('magic')
child.expect(">")

child.sendline('conf')
child.expect(">>")

child.sendline('system supplicant')
child.expect(">>")

child.sendline('set')
child.expect(" off | on ]:")

child.sendline('off')
child.expect(">>")

child.sendline('save')
child.expect(">>")

child.sendline("view")
child.expect(">>")

out = child.before
print(out)
#
child.expect("assword:")
child.sendline('alcatel')
# child.sendline('*<#/53#1/2')
print("logged in to host")

from time import sleep
from selenium import webdriver

import pexpect
import re

# driver = webdriverhttps://www.waketech.edu/programs-courses/credit/electrical-systems-technology/degrees-pathways.Chrome('/usr/local/bin/chromedriver')
driver = webdriver.Firefox(executable_path='/usr/local/bin/geckodriver')
# driver.get('http://www.google.com')
driver.get('http://192.168.1.254')
driver.implicitly_wait(20)
# driver.find_elements_by_tag_name("Settings") // this is for 599
driver.find_element_by_link_text("Settings").click()

# driver.findElement(By.linkText("Home Network")).click()
driver.implicitly_wait(20)

driver.find_element_by_link_text("LAN").click()
driver.implicitly_wait(20)
# driver.maximize_window()

driver.find_element_by_link_text("Wi-Fi").click()
driver.implicitly_wait(20)

password = driver.find_element_by_id("ADM_PASSWORD")

password.send_keys("8>1769&295")

driver.find_element_by_class_name('button').click()
# driver.find_element_by_xpath("//button[@value='Submit']").click()
# button.click()

driver.implicitly_wait(20)

# ENter device access code

sleep(30)
driver.quit()

child = pexpect.spawn("ssh root@192.168.1.254")
sleep(1)
# exit()


child.expect("assword:")
child.sendline('alcatel')
# child.sendline('*<#/53#1/2')
print("logged in to host")
child.expect("#")
child.sendline('exit')
sleep(5)
# driver.webdriver.quit()
exit()
# driver = webdriver.firefox('/home/palmer/.local/lib/python2.7/site-packages/chromedriver')
# driver = webdriver.chrome()
# browser.get('https://www.google.com')

#
# importing unicode_literals (above). spawnu accepts Unicode input and
# unicode_literals makes all string literals in this script Unicode by default.
p = pexpect.spawnu('uptime')

# This parses uptime output into the major groups using regex group matching.
p.expect(
    r'up\s+(.*?),\s+([0-9]+) users?,\s+load averages?: ([0-9]+\.[0-9][0-9]),?\s+([0-9]+\.[0-9][0-9]),?\s+([0-9]+\.[0-9][0-9])')
duration, users, av1, av5, av15 = p.match.groups()


days = '0'
hours = '0'
mins = '0'
if 'day' in duration:
    p.match = re.search(r'([0-9]+)\s+day', duration)
    days = str(int(p.match.group(1)))
if ':' in duration:
    p.match = re.search('([0-9]+):([0-9]+)', duration)
    hours = str(int(p.match.group(1)))
    mins = str(int(p.match.group(2)))
if 'min' in duration:
    p.match = re.search(r'([0-9]+)\s+min', duration)
    mins = str(int(p.match.group(1)))

# Print the parsed fields in CSV format.
# print('days, hours, minutes, users, cpu avg 1 min, cpu avg 5 min, cpu avg 15 min')
# print('%s, %s, %s, %s, %s, %s, %s' % (days, hours, mins, users, av1, av5, av15))


print("Turning off supplicant")
ip = "192.168.1.254"
password = '<#/53#1/2'
child = pexpect.spawn("telnet " "192.168.1.254")
sleep(1)
# exit()
child.expect("ogin:")
child.sendline("admin")

child.expect("assword:")
child.sendline('*<#/53#1/2')
child.expect(">")

child.sendline('magic')
child.expect(">")

child.sendline('conf')
child.expect(">>")

child.sendline('system supplicant')
child.expect(">>")

child.sendline('set')
child.expect(" off | on ]:")

child.sendline('off')
child.expect(">>")

child.sendline('save')
child.expect(">>")

child.sendline("view")
child.expect(">>")

out = child.before
print(out)
#

child.expect("#")
child.sendline('exit')
sleep(5)
# driver.webdriver.quit()
exit()
# driver = webdriver.firefox('/home/palmer/.local/lib/python2.7/site-packages/chromedriver')
# driver = webdriver.chrome()
# browser.get('https://www.google.com')



# Note that, for Python 3 compatibility reasons, we are using spawnu and
# importing unicode_literals (above). spawnu accepts Unicode input and
# unicode_literals makes all string literals in this script Unicode by default.
p = pexpect.spawnu('uptime')

# This parses uptime output into the major groups using regex group matching.
p.expect(
    r'up\s+(.*?),\s+([0-9]+) users?,\s+load averages?: ([0-9]+\.[0-9][0-9]),?\s+([0-9]+\.[0-9][0-9]),?\s+([0-9]+\.[0-9][0-9])')
duration, users, av1, av5, av15 = p.match.groups()

# The duration is a little harder to parse because of all the different
# styles of uptime. I'm sure there is a way to do this all at once with
# one single regex, but I bet it would be hard to read and maintain.
# If anyone wants to send me a version using a single regex I'd be happy to see it.
days = '0'
hours = '0'
mins = '0'
if 'day' in duration:
    p.match = re.search(r'([0-9]+)\s+day', duration)
    days = str(int(p.match.group(1)))
if ':' in duration:
    p.match = re.search('([0-9]+):([0-9]+)', duration)
    hours = str(int(p.match.group(1)))
    mins = str(int(p.match.group(2)))
if 'min' in duration:
    p.match = re.search(r'([0-9]+)\s+min', duration)
    mins = str(int(p.match.group(1)))

